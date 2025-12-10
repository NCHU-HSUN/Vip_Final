import numpy as np
from typing import IO, Iterator, Tuple, List, Dict, Optional, Callable, Any
import itertools
import os
import sys
import math
import time
from datetime import datetime
import concurrent.futures

# --- 1. 匯入 Numba JIT ---
try:
    from numba import jit, int32, int64
    from numba.experimental import jitclass
except ImportError:
    print("警告: 'numba' 未安裝。程式將以純 Python 慢速模式執行。")
    def jit(nopython=True, nogil=True):
        def decorator(func): return func
        return decorator
    def jitclass(cls_or_spec=None, spec=None):
        def decorator(cls): return cls
        return decorator
    int32 = int
    int64 = int

# --- 2. 匯入輔助套件 ---
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None

try:
    from PIL import Image
except ImportError:
    print("警告: 未安裝 Pillow，無法輸出驗證圖片。")
    Image = None

# ==========================================
# 核心類別
# ==========================================
class Frame:
    def __init__(self, y: np.ndarray, cb: np.ndarray, cr: np.ndarray):
        self.y, self.cb, self.cr = y, cb, cr
        self.height, self.width = y.shape
        
    @staticmethod
    def create_blank(width, height):
        return Frame(np.full((height, width), 128, dtype=np.uint8),
                     np.full((height//2, width//2), 128, dtype=np.uint8),
                     np.full((height//2, width//2), 128, dtype=np.uint8))
                     
    def put_y_macroblock(self, r, c, block):
        self.y[r:r+16, c:c+16] = block

class Y4MReader:
    def __init__(self, filepath: str):
        if not os.path.exists(filepath):
            raise FileNotFoundError(f"找不到影片檔案: {filepath}")
        self.file = open(filepath, 'rb')
        self._parse_header()

    def _parse_header(self):
        line = self.file.readline().decode('ascii').strip()
        for p in line.split(' '):
            if p.startswith('W'): self.width = int(p[1:])
            elif p.startswith('H'): self.height = int(p[1:])
        self.y_size = self.width * self.height
        self.c_size = (self.width // 2) * (self.height // 2)

    def __iter__(self): return self

    def __next__(self) -> Frame:
        if not self.file.readline(): raise StopIteration
        y = np.frombuffer(self.file.read(self.y_size), dtype=np.uint8).reshape((self.height, self.width))
        c_sz = self.c_size
        cb = np.frombuffer(self.file.read(c_sz), dtype=np.uint8).reshape((self.height//2, self.width//2))
        cr = np.frombuffer(self.file.read(c_sz), dtype=np.uint8).reshape((self.height//2, self.width//2))
        return Frame(y, cb, cr)

    def close(self): self.file.close()

def pad_frame(frame: Frame) -> Frame:
    h, w = frame.y.shape
    ph, pw = (16 - h % 16) % 16, (16 - w % 16) % 16
    if ph == 0 and pw == 0: return frame
    return Frame(np.pad(frame.y, ((0, ph), (0, pw)), 'edge'),
                 np.pad(frame.cb, ((0, ph//2), (0, pw//2)), 'edge'),
                 np.pad(frame.cr, ((0, ph//2), (0, pw//2)), 'edge'))

# ==========================================
# JIT 函式
# ==========================================
spec = [('mv_r', int32), ('mv_c', int32), ('sad', int64), ('check_points', int32)]
@jitclass(spec)
class MotionVectorResult:
    def __init__(self, mv_r, mv_c, sad, check_points):
        self.mv_r, self.mv_c, self.sad, self.check_points = mv_r, mv_c, sad, check_points

@jit(nopython=True, nogil=True) 
def calculate_sad(b1: np.ndarray, b2: np.ndarray) -> int:
    return np.sum(np.abs(b1.astype(np.int16) - b2.astype(np.int16)))

@jit(nopython=True, nogil=True) 
def get_ref_block(ref: np.ndarray, r: int, c: int, dr: int, dc: int) -> np.ndarray:
    h, w = ref.shape
    rr = min(max(r + dr, 0), h - 16)
    cc = min(max(c + dc, 0), w - 16)
    return ref[rr:rr+16, cc:cc+16]

def calc_psnr(orig: Frame, recon: Frame) -> float:
    mse = np.mean((orig.y.astype(np.float32) - recon.y.astype(np.float32)) ** 2)
    return 20 * math.log10(255.0) - 10 * math.log10(mse) if mse > 0 else float('inf')

# ==========================================
# 演算法 (Full, TSS, Diamond, HEXBS)
# ==========================================
@jit(nopython=True, nogil=True)
def algo_full_search(cur, ref, r, c, rng):
    best_sad, best_dr, best_dc, checks = 999999, 0, 0, 0
    for dr in range(-rng, rng + 1):
        for dc in range(-rng, rng + 1):
            sad = calculate_sad(cur, get_ref_block(ref, r, c, dr, dc))
            checks += 1
            if sad < best_sad: best_sad, best_dr, best_dc = sad, dr, dc
    return MotionVectorResult(best_dr, best_dc, best_sad, checks)

@jit(nopython=True, nogil=True)
def algo_tss(cur, ref, r, c, rng):
    best_sad, best_dr, best_dc, checks, step = 999999, 0, 0, 0, 4
    neighbors = [(-1, -1), (-1, 0), (-1, 1), (0, -1), (0, 1), (1, -1), (1, 0), (1, 1)]
    while step >= 1:
        check_center = (step == 4) 
        ls, lr, lc = best_sad, best_dr, best_dc
        if check_center:
            dr, dc = best_dr, best_dc
            if -rng <= dr <= rng and -rng <= dc <= rng:
                sad = calculate_sad(cur, get_ref_block(ref, r, c, dr, dc))
                checks += 1
                if sad < ls: ls, lr, lc = sad, dr, dc
        for i, j in neighbors:
            dr, dc = best_dr + i * step, best_dc + j * step
            if not (-rng <= dr <= rng and -rng <= dc <= rng): continue
            sad = calculate_sad(cur, get_ref_block(ref, r, c, dr, dc))
            checks += 1
            if sad < ls: ls, lr, lc = sad, dr, dc
        best_sad, best_dr, best_dc, step = ls, lr, lc, step // 2
    return MotionVectorResult(best_dr, best_dc, best_sad, checks)

@jit(nopython=True, nogil=True)
def algo_diamond(cur, ref, r, c, rng):
    LDSP = [(0,0), (0,2), (0,-2), (2,0), (-2,0), (1,1), (1,-1), (-1,1), (-1,-1)]
    SDSP = [(0,0), (0,1), (0,-1), (1,0), (-1,0)]
    cache = np.zeros((2*rng+1, 2*rng+1), dtype=np.int8)
    checks, best_dr, best_dc = 0, 0, 0
    best_sad = calculate_sad(cur, get_ref_block(ref, r, c, 0, 0))
    cache[rng, rng] = 1; checks += 1
    
    while True:
        ctr_r, ctr_c, found = best_dr, best_dc, False
        for i, j in LDSP:
            dr, dc = ctr_r + i, ctr_c + j
            if not (-rng<=dr<=rng and -rng<=dc<=rng) or cache[dr+rng, dc+rng]: continue
            sad = calculate_sad(cur, get_ref_block(ref, r, c, dr, dc))
            checks += 1; cache[dr+rng, dc+rng] = 1
            if sad < best_sad: best_sad, best_dr, best_dc, found = sad, dr, dc, True
        if not found: break
    for i, j in SDSP:
        dr, dc = best_dr + i, best_dc + j
        if not (-rng<=dr<=rng and -rng<=dc<=rng) or cache[dr+rng, dc+rng]: continue
        sad = calculate_sad(cur, get_ref_block(ref, r, c, dr, dc))
        checks += 1; cache[dr+rng, dc+rng] = 1
        if sad < best_sad: best_sad, best_dr, best_dc = sad, dr, dc
    return MotionVectorResult(best_dr, best_dc, best_sad, checks)

@jit(nopython=True, nogil=True)
def algo_hexbs(cur, ref, r, c, rng):
    LHP = [(2,0), (1,2), (-1,2), (-2,0), (-1,-2), (1,-2)]
    SHP = [(1,0), (0,1), (-1,0), (0,-1)]
    cache = np.zeros((2*rng+1, 2*rng+1), dtype=np.int8)
    checks, best_dr, best_dc = 0, 0, 0
    best_sad = calculate_sad(cur, get_ref_block(ref, r, c, 0, 0))
    cache[rng, rng] = 1; checks += 1
    
    while True:
        ctr_r, ctr_c, center_best = best_dr, best_dc, True
        for i, j in LHP:
            dr, dc = ctr_r + i, ctr_c + j
            if not (-rng<=dr<=rng and -rng<=dc<=rng) or cache[dr+rng, dc+rng]: continue
            sad = calculate_sad(cur, get_ref_block(ref, r, c, dr, dc))
            checks += 1; cache[dr+rng, dc+rng] = 1
            if sad < best_sad: best_sad, best_dr, best_dc, center_best = sad, dr, dc, False
        if center_best: break
    for i, j in SHP:
        dr, dc = best_dr + i, best_dc + j
        if not (-rng<=dr<=rng and -rng<=dc<=rng) or cache[dr+rng, dc+rng]: continue
        sad = calculate_sad(cur, get_ref_block(ref, r, c, dr, dc))
        checks += 1; cache[dr+rng, dc+rng] = 1
        if sad < best_sad: best_sad, best_dr, best_dc = sad, dr, dc
    return MotionVectorResult(best_dr, best_dc, best_sad, checks)

# ==========================================
# 分析流程
# ==========================================
def process_row_task(row_idx, width, cur_y, ref_y, algo_func, rng):
    results = []
    for col_idx in range(0, width, 16):
        cur_mb = cur_y[row_idx : row_idx+16, col_idx : col_idx+16].copy()
        res = algo_func(cur_mb, ref_y, row_idx, col_idx, rng)
        results.append(res)
    return row_idx, results

def save_verification_images(vname, algo_name, cur_frame, recon_frame, out_dir):
    if Image is None: return
    try:
        recon_img = Image.fromarray(recon_frame.y)
        recon_img.save(f"{out_dir}/{vname}_{algo_name}_Recon.png")
        
        diff = np.abs(cur_frame.y.astype(int) - recon_frame.y.astype(int)).astype(np.uint8)
        diff_enhanced = diff * 5 
        diff_img = Image.fromarray(diff_enhanced)
        diff_img.save(f"{out_dir}/{vname}_{algo_name}_Error.png")
    except Exception as e:
        print(f"  [Error] 儲存圖片失敗: {e}")

def analyze_video(name, path, algos, rng, executor, save_images=False, out_dir=""):
    stats = {k: {'psnr':[], 'time':[], 'checks':[]} for k in algos}
    
    # 詳細數據記錄 (Frame-by-Frame)
    # 格式: [Algo, Frame_ID, PSNR, Checks, Time]
    detailed_data = []  
    
    try:
        reader = Y4MReader(path)
    except Exception as e:
        print(f"  [Error] 無法讀取影片 {path}: {e}")
        return {}, []

    try:
        ref = pad_frame(next(reader))
    except StopIteration: return {}, []
        
    try:
        frame_idx = 0
        while True:
            try: cur = pad_frame(next(reader))
            except StopIteration: break
            
            for algo_name, algo_func in algos.items():
                t_start = time.perf_counter()
                futures = []
                for r in range(0, cur.height, 16):
                    futures.append(executor.submit(process_row_task, r, cur.width, cur.y, ref.y, algo_func, rng))
                
                total_checks = 0
                recon = Frame.create_blank(cur.width, cur.height)
                for f in concurrent.futures.as_completed(futures):
                    r_idx, row_res = f.result()
                    for i, mb_res in enumerate(row_res):
                        c_idx = i * 16
                        total_checks += mb_res.check_points
                        ref_blk = get_ref_block(ref.y, r_idx, c_idx, mb_res.mv_r, mb_res.mv_c)
                        recon.put_y_macroblock(r_idx, c_idx, ref_blk)
                
                t_end = time.perf_counter()
                mb_count = (cur.height // 16) * (cur.width // 16)
                
                frame_psnr = calc_psnr(cur, recon)
                frame_checks = total_checks / mb_count
                frame_time = t_end - t_start
                
                stats[algo_name]['time'].append(frame_time)
                stats[algo_name]['checks'].append(frame_checks)
                stats[algo_name]['psnr'].append(frame_psnr)
                
                detailed_data.append([algo_name, frame_idx, frame_psnr, frame_checks, frame_time])
                
                if save_images and frame_idx == 0:
                    save_verification_images(name, algo_name, cur, recon, out_dir)
            
            ref = cur
            frame_idx += 1
    finally:
        reader.close()
    return stats, detailed_data

# ==========================================
# 繪圖與報表 (超級升級版: 分欄位儲存)
# ==========================================
def save_excel_report_detailed(final_avg, frame_stats_avg, output_path):
    if not openpyxl: 
        print("未安裝 openpyxl，無法輸出 Excel。")
        return
        
    wb = openpyxl.Workbook()
    
    # --- 1. Summary Sheet (總表) ---
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum.append(["Video", "Algorithm", "Avg PSNR (dB)", "Avg Checks", "Avg Time (s)", "PSNR Diff", "Checks Red.%", "Speedup"])
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", fill_type="solid")
    for cell in ws_sum[1]: cell.font = header_font; cell.fill = header_fill

    for vname, algos_data in final_avg.items():
        fs_psnr, fs_checks, fs_time = None, None, None
        if "FullSearch" in algos_data:
            d = algos_data["FullSearch"]
            fs_psnr, fs_checks, fs_time = np.mean(d['psnr']), np.mean(d['checks']), np.mean(d['time'])
        
        for algo, d in algos_data.items():
            avg_p, avg_c, avg_t = np.mean(d['psnr']), np.mean(d['checks']), np.mean(d['time'])
            p_diff, c_red, spd = "-", "-", "-"
            if fs_psnr is not None and fs_checks is not None and fs_time is not None:
                p_diff = f"{avg_p - fs_psnr:+.3f}"
                c_red = f"{(fs_checks - avg_c)/fs_checks*100:.1f}%"
                spd = f"{fs_time/avg_t:.1f}x" if avg_t > 0 else "-"
            ws_sum.append([vname, algo, round(avg_p,2), round(avg_c,1), round(avg_t,5), p_diff, c_red, spd])

    # --- 2. Detailed Sheets (每個影片一個 Sheet，欄位並排) ---
    # frame_stats_avg 結構: {(Video, Algo, Frame) -> (AvgPSNR, AvgChecks, AvgTime)}
    
    # 先整理出有哪些影片和演算法
    videos = sorted(list(set(k[0] for k in frame_stats_avg.keys())))
    algos_set = sorted(list(set(k[1] for k in frame_stats_avg.keys())))
    # 確保 FullSearch 排在第一個，方便比較
    if "FullSearch" in algos_set:
        algos_set.remove("FullSearch")
        algos_set.insert(0, "FullSearch")

    for vname in videos:
        # 建立新的 Sheet，名稱為影片名
        ws = wb.create_sheet(title=f"{vname}_Details")
        
        # 製作標題列: Frame | Full_PSNR | Full_Checks | Full_Time | TSS_PSNR ...
        headers = ["Frame Index"]
        for algo in algos_set:
            headers.extend([f"{algo}_PSNR", f"{algo}_Checks", f"{algo}_Time"])
        ws.append(headers)
        
        # 標題美化
        for cell in ws[1]: cell.font = header_font; cell.fill = header_fill
        
        # 找出該影片最大的 Frame Index
        max_frame = max([k[2] for k in frame_stats_avg.keys() if k[0] == vname])
        
        # 逐列填入數據
        for f_idx in range(max_frame + 1):
            row_data = [f_idx]
            for algo in algos_set:
                key = (vname, algo, f_idx)
                if key in frame_stats_avg:
                    p, c, t = frame_stats_avg[key]
                    row_data.extend([round(p, 4), round(c, 1), round(t, 6)])
                else:
                    row_data.extend(["-", "-", "-"])
            ws.append(row_data)
            
        # 自動調整欄寬
        for column in ws.columns:
            max_length = 0
            col_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
                except: pass
            ws.column_dimensions[col_letter].width = max_length + 2

    # 自動調整 Summary 欄寬
    for column in ws_sum.columns:
        max_length = 0
        col_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
            except: pass
        ws_sum.column_dimensions[col_letter].width = max_length + 2

    wb.save(output_path)
    print(f"\n✅ 詳細 Excel 報告已儲存: {output_path}")

def plot_charts(video_name, stats, out_dir):
    import matplotlib.pyplot as plt
    fig, (ax1, ax2, ax3) = plt.subplots(3, 1, figsize=(10, 16))
    ax1.set_title(f'{video_name} - PSNR'); ax1.set_ylabel('dB')
    ax2.set_title(f'{video_name} - Checks'); ax2.set_ylabel('Count')
    ax3.set_title(f'{video_name} - Time'); ax3.set_ylabel('Sec')
    
    fs_txt = "Full Search (Benchmark):\n"
    has_fs = False
    for name, data in stats.items():
        x = range(len(data['time']))
        if 'Full' in name:
            has_fs = True
            fs_txt += f" Checks:{np.mean(data['checks']):.1f}, Time:{np.mean(data['time']):.4f}s"
            ax1.plot(x, data['psnr'], 'k--', alpha=0.5, label=name)
        else:
            ax1.plot(x, data['psnr'], label=name); ax2.plot(x, data['checks'], label=name); ax3.plot(x, data['time'], label=name)
            
    for ax in [ax1, ax2, ax3]: ax.legend(); ax.grid(True)
    if has_fs: fig.text(0.5, 0.02, fs_txt, ha='center', bbox=dict(fc='white', alpha=0.9))
    fig.savefig(f"{out_dir}/{video_name}_analysis.png")
    plt.close(fig)

if __name__ == '__main__':
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_dir = f"ME_Run_{ts}"
    os.makedirs(out_dir, exist_ok=True)
    os.makedirs(f"{out_dir}/plots", exist_ok=True)
    
    base_video_path = "video"
    if not os.path.exists(base_video_path):
        possible_paths = ["../video", "./video", "nchu-hsun/vip_final/NCHU-HSUN-Vip_Final-84430069098edb7f242cd73eaae9c798ab3ce302/video"]
        for p in possible_paths:
            if os.path.exists(p):
                base_video_path = p
                print(f"📍 偵測到影片資料夾位置: {base_video_path}")
                break
    
    VIDEOS = {
        "Garden": os.path.join(base_video_path, "garden_sif.y4m"), 
        "Football": os.path.join(base_video_path, "football_cif.y4m"), 
        "Tennis": os.path.join(base_video_path, "tennis_sif.y4m")
    }
    
    VALID_VIDEOS = {}
    for k, v in VIDEOS.items():
        if os.path.exists(v): VALID_VIDEOS[k] = v
        else: print(f"⚠️ 警告: 找不到檔案 {v}，將跳過此測試。")
            
    ALGOS = {"FullSearch": algo_full_search, "TSS": algo_tss, "Diamond": algo_diamond, "HEXBS": algo_hexbs}
    
    workers = os.cpu_count() or 4
    NUM_RUNS = 3 # 恢復為 10 次，讓平均值更有意義
    
    print(f"=== 啟動驗證與效能測試 (Workers: {workers}) ===")
    print(f"📁 結果將儲存於: {out_dir}")
    
    final_avg = {}
    # 用於累積 Frame-Level 的詳細數據: Key=(Video, Algo, Frame), Value={'psnr':[], ...}
    frame_stats_accumulator = {} 

    with concurrent.futures.ThreadPoolExecutor(max_workers=workers) as executor:
        for run_i in range(NUM_RUNS):
            print(f"\n📢 Run {run_i + 1}/{NUM_RUNS} ...")
            save_imgs = (run_i == 0)
            
            for vname, vpath in VALID_VIDEOS.items():
                print(f"   -> {vname}...", end='', flush=True)
                
                cur_res, cur_details = analyze_video(vname, vpath, ALGOS, 15, executor, save_images=save_imgs, out_dir=out_dir)
                print(" Done")
                
                # 1. 累積 Summary 數據
                if vname not in final_avg:
                    final_avg[vname] = {algo: {'psnr': d['psnr'], 'checks': d['checks'], 'time': d['time']} for algo, d in cur_res.items()}
                else:
                    for algo in ALGOS:
                        final_avg[vname][algo]['time'] = [sum(x) for x in zip(final_avg[vname][algo]['time'], cur_res[algo]['time'])]
                        
                # 2. 累積 Detailed Frame 數據 (為了計算平均)
                for row in cur_details:
                    # row = [Algo, Frame, PSNR, Checks, Time]
                    algo, fr, p, c, t = row
                    key = (vname, algo, fr)
                    if key not in frame_stats_accumulator:
                        frame_stats_accumulator[key] = {'psnr':[], 'checks':[], 'time':[]}
                    frame_stats_accumulator[key]['psnr'].append(p)
                    frame_stats_accumulator[key]['checks'].append(c)
                    frame_stats_accumulator[key]['time'].append(t)

    if not final_avg:
        print("\n❌ 沒有任何影片被成功分析，請檢查路徑。")
    else:
        print("\n📊 計算平均值並產出報告...")
        
        # 1. 處理 Summary 平均
        for vname in final_avg:
            for algo in final_avg[vname]:
                final_avg[vname][algo]['time'] = [t/NUM_RUNS for t in final_avg[vname][algo]['time']]
            plot_charts(vname, final_avg[vname], f"{out_dir}/plots")
            
        # 2. 處理 Detailed Frame 平均
        frame_stats_avg = {}
        for key, vals in frame_stats_accumulator.items():
            avg_p = np.mean(vals['psnr'])
            avg_c = np.mean(vals['checks'])
            avg_t = np.mean(vals['time'])
            frame_stats_avg[key] = (avg_p, avg_c, avg_t)

        save_excel_report_detailed(final_avg, frame_stats_avg, f"{out_dir}/ME_Result_Detailed_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        print("\n=== 全部完成，請檢查資料夾中的圖表與 Excel ===")