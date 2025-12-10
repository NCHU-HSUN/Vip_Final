import numpy as np
from typing import IO, Iterator, Tuple, List, Dict, Optional, Callable, Any
import itertools
import os
import sys
import math
import time
from datetime import datetime
import concurrent.futures

# --- [設定] 實驗參數設定區 ---
# 您可以在這裡修改循環次數，設為 3~5 次可以得到更穩定的時間數據
TEST_LOOPS = 3  
SEARCH_RANGE = 15
MAX_FRAMES = None  # None 表示讀取整部影片，或是設為整數 (例如 50) 來限制測試幀數

# --- [優化] 匯入 Numba JIT ---
try:
    from numba import jit
    from numba.experimental import jitclass
    from numba import int32, int64
except ImportError:
    print("警告: 'numba' 函式庫未找到。")
    
    def jit(**kwargs):
        def decorator(func):
            return func
        return decorator

    def jitclass(cls_or_spec=None, spec=None):
        def decorator(cls):
            return cls
        if callable(cls_or_spec):
            return decorator(cls_or_spec)
        return decorator

    int32 = int
    int64 = int

# --- [新功能] 匯入 openpyxl ---
try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None

# ==========================================
# 核心類別與讀取器
# ==========================================

class Frame:
    def __init__(self, y: np.ndarray, cb: np.ndarray, cr: np.ndarray):
        self.y = y
        self.cb = cb
        self.cr = cr
        self.height, self.width = y.shape
        
    def iter_y_macroblocks(self) -> Iterator[Tuple[int, int, np.ndarray]]:
        mb_size = 16
        for r in range(0, self.height, mb_size):
            for c in range(0, self.width, mb_size):
                y_block = self.y[r:r+mb_size, c:c+mb_size]
                yield r, c, y_block

    @staticmethod
    def create_blank(width, height):
        y = np.full((height, width), 128, dtype=np.uint8)
        cb = np.full((height // 2, width // 2), 128, dtype=np.uint8)
        cr = np.full((height // 2, width // 2), 128, dtype=np.uint8)
        return Frame(y, cb, cr)

    def put_y_macroblock(self, r: int, c: int, y_block_16x16: np.ndarray):
        self.y[r:r+16, c:c+16] = y_block_16x16

class Y4MReader:
    def __init__(self, filepath: str):
        self.file = open(filepath, 'rb')
        self.width = 0
        self.height = 0
        self.y_size = 0
        self.c_size = 0
        self._parse_header()

    def _parse_header(self):
        header_line = self.file.readline().decode('ascii').strip()
        if not header_line.startswith('YUV4MPEG2'):
            raise ValueError("檔案不是有效的 YUV4MPEG2 格式")
        
        params = header_line.split(' ')
        for p in params:
            if p.startswith('W'):
                self.width = int(p[1:])
            elif p.startswith('H'):
                self.height = int(p[1:])
        
        if self.width == 0 or self.height == 0:
            raise ValueError("未能在 Y4M 標頭中找到寬度或高度")
            
        self.y_size = self.width * self.height
        self.c_size = (self.width // 2) * (self.height // 2)

    def __iter__(self) -> Iterator[Frame]:
        return self

    def __next__(self) -> Frame:
        frame_header = self.file.readline()
        if not frame_header:
            raise StopIteration
            
        if not frame_header.startswith(b'FRAME'):
            raise IOError(f"預期的 FRAME 標頭未找到")
        
        try:
            y_data = self.file.read(self.y_size)
            cb_data = self.file.read(self.c_size)
            cr_data = self.file.read(self.c_size)
            
            if len(y_data) != self.y_size:
                raise StopIteration
                
            y = np.frombuffer(y_data, dtype=np.uint8).reshape((self.height, self.width))
            cb = np.frombuffer(cb_data, dtype=np.uint8).reshape((self.height // 2, self.width // 2))
            cr = np.frombuffer(cr_data, dtype=np.uint8).reshape((self.height // 2, self.width // 2))
            
            return Frame(y, cb, cr)
        except Exception as e:
            print(f"讀取錯誤: {e}")
            raise StopIteration

    def close(self):
        self.file.close()

def pad_frame_to_macroblock_boundary(frame: Frame) -> Frame:
    h, w = frame.y.shape
    mh, mw = 16, 16
    pad_h = (mh - h % mh) % mh
    pad_w = (mw - w % mw) % mw
    if pad_h == 0 and pad_w == 0:
        return frame
    y_padded = np.pad(frame.y, ((0, pad_h), (0, pad_w)), mode='edge')
    cb_padded = np.pad(frame.cb, ((0, pad_h // 2), (0, pad_w // 2)), mode='edge')
    cr_padded = np.pad(frame.cr, ((0, pad_h // 2), (0, pad_w // 2)), mode='edge')
    return Frame(y_padded, cb_padded, cr_padded)

# ==========================================
# JIT 加速函式
# ==========================================

spec = [
    ('mv_r', int32),
    ('mv_c', int32),
    ('sad', int64),
    ('check_points', int32),
]

@jitclass(spec)
class MotionVectorResult:
    def __init__(self, mv_r: int, mv_c: int, sad: int, check_points: int):
        self.mv_r = mv_r
        self.mv_c = mv_c
        self.sad = sad
        self.check_points = check_points

@jit(nopython=True, nogil=True) 
def calculate_sad(block1: np.ndarray, block2: np.ndarray) -> np.int64:
    diff = block1.astype(np.int16) - block2.astype(np.int16)
    return np.sum(np.abs(diff))

@jit(nopython=True, nogil=True) 
def get_reference_block(ref_y: np.ndarray, mb_r: int, mb_c: int, mv_r: int, mv_c: int) -> np.ndarray:
    h, w = ref_y.shape
    mb_size = 16
    ref_r_start = mb_r + mv_r
    ref_c_start = mb_c + mv_c
    r_max = h - mb_size
    ref_r = 0 if ref_r_start < 0 else (r_max if ref_r_start > r_max else ref_r_start)
    c_max = w - mb_size
    ref_c = 0 if ref_c_start < 0 else (c_max if ref_c_start > c_max else ref_c_start)
    return ref_y[ref_r : ref_r + mb_size, ref_c : ref_c + mb_size]

def calculate_psnr(frame_orig: Frame, frame_recon: Frame) -> float:
    mse = np.mean((frame_orig.y.astype(np.float32) - frame_recon.y.astype(np.float32)) ** 2)
    if mse == 0: return float('inf')
    return 20 * math.log10(255.0) - 10 * math.log10(mse)

# ==========================================
# ME 演算法 (Full, TSS, Diamond, HEXBS)
# ==========================================

@jit(nopython=True, nogil=True)
def full_search(current_mb_y: np.ndarray, ref_frame_y: np.ndarray, mb_r: int, mb_c: int, search_range: int = 7):
    best_sad = np.int64(999999999999)
    best_mv_r, best_mv_c = 0, 0
    check_points = 0
    for r_vec in range(-search_range, search_range + 1):
        for c_vec in range(-search_range, search_range + 1):
            ref_block = get_reference_block(ref_frame_y, mb_r, mb_c, r_vec, c_vec)
            sad = calculate_sad(current_mb_y, ref_block)
            check_points += 1
            if sad < best_sad:
                best_sad = sad
                best_mv_r, best_mv_c = r_vec, c_vec
    return MotionVectorResult(best_mv_r, best_mv_c, best_sad, check_points)

@jit(nopython=True, nogil=True)
def three_step_search(current_mb_y: np.ndarray, ref_frame_y: np.ndarray, mb_r: int, mb_c: int, search_range: int = 7):
    check_points = 0
    steps = np.array([4, 2, 1]) 
    center_r, center_c = 0, 0
    best_sad = np.int64(999999999999)
    best_mv_r, best_mv_c = 0, 0
    offsets = np.array([[0, 0], [0, 1], [0, -1], [1, 0], [-1, 0], [1, 1], [1, -1], [-1, 1], [-1, -1]])
    
    for i in range(steps.shape[0]):
        step = steps[i]
        step_best_sad = np.int64(999999999999)
        step_best_r, step_best_c = center_r, center_c
        points = offsets if i == 0 else offsets[1:]
        for j in range(points.shape[0]):
            mv_r = center_r + (points[j, 0] * step)
            mv_c = center_c + (points[j, 1] * step)
            if not (-search_range <= mv_r <= search_range and -search_range <= mv_c <= search_range): continue
            ref_block = get_reference_block(ref_frame_y, mb_r, mb_c, mv_r, mv_c)
            sad = calculate_sad(current_mb_y, ref_block)
            check_points += 1
            if sad < step_best_sad:
                step_best_sad = sad
                step_best_r, step_best_c = mv_r, mv_c
        if step_best_sad < best_sad:
            best_sad = step_best_sad
            best_mv_r, best_mv_c = step_best_r, step_best_c
        center_r, center_c = step_best_r, step_best_c
    return MotionVectorResult(best_mv_r, best_mv_c, best_sad, check_points)

@jit(nopython=True, nogil=True)
def diamond_search(current_mb_y: np.ndarray, ref_frame_y: np.ndarray, mb_r: int, mb_c: int, search_range: int = 7):
    LDSP = [(0, 0), (0, 2), (0, -2), (2, 0), (-2, 0), (1, 1), (1, -1), (-1, 1), (-1, -1)]
    SDSP = [(0, 0), (0, 1), (0, -1), (1, 0), (-1, 0)]
    check_points = 0
    cache_size = 2 * search_range + 1
    checked_cache = np.zeros((cache_size, cache_size), dtype=np.int8)
    
    best_r, best_c = 0, 0
    ref_blk = get_reference_block(ref_frame_y, mb_r, mb_c, 0, 0)
    best_sad = calculate_sad(current_mb_y, ref_blk)
    checked_cache[0 + search_range, 0 + search_range] = 1
    check_points += 1
    
    while True:
        curr_center_r, curr_center_c = best_r, best_c
        found_better = False
        for dr, dc in LDSP:
            r = curr_center_r + dr
            c = curr_center_c + dc
            if not (-search_range <= r <= search_range and -search_range <= c <= search_range): continue
            if checked_cache[r + search_range, c + search_range] == 1: continue
            sad = calculate_sad(current_mb_y, get_reference_block(ref_frame_y, mb_r, mb_c, r, c))
            check_points += 1
            checked_cache[r + search_range, c + search_range] = 1
            if sad < best_sad:
                best_sad = sad
                best_r, best_c = r, c
                found_better = True
        if not found_better or (best_r == curr_center_r and best_c == curr_center_c):
            break

    for dr, dc in SDSP:
        r = best_r + dr
        c = best_c + dc
        if not (-search_range <= r <= search_range and -search_range <= c <= search_range): continue
        if checked_cache[r + search_range, c + search_range] == 1: continue
        sad = calculate_sad(current_mb_y, get_reference_block(ref_frame_y, mb_r, mb_c, r, c))
        check_points += 1
        checked_cache[r + search_range, c + search_range] = 1
        if sad < best_sad:
            best_sad = sad
            best_r, best_c = r, c
    return MotionVectorResult(best_r, best_c, best_sad, check_points)

@jit(nopython=True, nogil=True)
def hexagon_search(current_mb_y: np.ndarray, ref_frame_y: np.ndarray, mb_r: int, mb_c: int, search_range: int = 7):
    LHP_OFFSETS = [(2, 0), (1, 2), (-1, 2), (-2, 0), (-1, -2), (1, -2)]
    SHP_OFFSETS = [(1, 0), (0, 1), (-1, 0), (0, -1)]
    check_points = 0
    cache_size = 2 * search_range + 1
    checked_cache = np.zeros((cache_size, cache_size), dtype=np.int8)
    
    center_r, center_c = 0, 0
    ref_block = get_reference_block(ref_frame_y, mb_r, mb_c, center_r, center_c)
    best_sad = calculate_sad(current_mb_y, ref_block)
    best_mv_r, best_mv_c = center_r, center_c
    checked_cache[center_r + search_range, center_c + search_range] = 1
    check_points += 1
    
    while True:
        current_center_r, current_center_c = best_mv_r, best_mv_c
        best_point_is_center = True
        for dr, dc in LHP_OFFSETS:
            mv_r = current_center_r + dr
            mv_c = current_center_c + dc
            if not (-search_range <= mv_r <= search_range and -search_range <= mv_c <= search_range): continue
            if checked_cache[mv_r + search_range, mv_c + search_range] == 1: continue
            ref_block = get_reference_block(ref_frame_y, mb_r, mb_c, mv_r, mv_c)
            sad = calculate_sad(current_mb_y, ref_block)
            check_points += 1
            checked_cache[mv_r + search_range, mv_c + search_range] = 1
            if sad < best_sad:
                best_sad = sad
                best_mv_r, best_mv_c = mv_r, mv_c
                best_point_is_center = False
        if best_point_is_center: break
            
    for dr, dc in SHP_OFFSETS:
        mv_r = best_mv_r + dr
        mv_c = best_mv_c + dc
        if not (-search_range <= mv_r <= search_range and -search_range <= mv_c <= search_range): continue
        if checked_cache[mv_r + search_range, mv_c + search_range] == 1: continue
        ref_block = get_reference_block(ref_frame_y, mb_r, mb_c, mv_r, mv_c)
        sad = calculate_sad(current_mb_y, ref_block)
        check_points += 1
        checked_cache[mv_r + search_range, mv_c + search_range] = 1
        if sad < best_sad:
            best_sad = sad
            best_mv_r, best_mv_c = mv_r, mv_c
    return MotionVectorResult(best_mv_r, best_mv_c, best_sad, check_points)

# ==========================================
# 分析流程
# ==========================================

def reconstruct_frame_from_mvs(ref_frame: Frame, mv_results: List[Tuple[int, int, MotionVectorResult]]) -> Frame:
    recon_frame = Frame.create_blank(ref_frame.width, ref_frame.height)
    for r, c, mv_res in mv_results:
        ref_block_y = get_reference_block(ref_frame.y, r, c, mv_res.mv_r, mv_res.mv_c)
        recon_frame.put_y_macroblock(r, c, ref_block_y)
    return recon_frame

def process_row_of_blocks(row_idx: int, width: int, current_y: np.ndarray, ref_y: np.ndarray, algo_func: Callable, search_range: int):
    mb_size = 16
    row_results = []
    for c in range(0, width, mb_size):
        current_mb = current_y[row_idx:row_idx+mb_size, c:c+mb_size].copy()
        mv_result = algo_func(current_mb, ref_y, row_idx, c, search_range)
        row_results.append((row_idx, c, mv_result))
    return row_results

def analyze_sequence(frame_iterator: Iterator[Frame], algorithm_func: Callable, search_range: int, executor: concurrent.futures.ThreadPoolExecutor) -> Dict[str, Any]:
    psnr_curve, avg_check_points_curve, processing_time_curve = [], [], []
    processed_frame_pairs = 0
    
    try:
        ref_frame = pad_frame_to_macroblock_boundary(next(frame_iterator))
        
        while True:
            current_frame = pad_frame_to_macroblock_boundary(next(frame_iterator))
            start_time = time.perf_counter()
            
            mv_results_for_frame = []
            total_checks = 0
            futures = []
            mb_size = 16
            height, width = current_frame.height, current_frame.width
            
            # 使用傳入的 executor
            for r in range(0, height, mb_size):
                future = executor.submit(
                    process_row_of_blocks, 
                    r, width, current_frame.y, ref_frame.y, algorithm_func, search_range
                )
                futures.append(future)
            
            for future in concurrent.futures.as_completed(futures):
                row_results = future.result()
                for r, c, mv_res in row_results:
                    mv_results_for_frame.append((r, c, mv_res))
                    total_checks += mv_res.check_points

            if not mv_results_for_frame: continue
            
            recon_frame = reconstruct_frame_from_mvs(ref_frame, mv_results_for_frame)
            end_time = time.perf_counter()

            psnr_curve.append(calculate_psnr(current_frame, recon_frame))
            avg_check_points_curve.append(total_checks / len(mv_results_for_frame))
            processing_time_curve.append(end_time - start_time)
            processed_frame_pairs += 1
            ref_frame = current_frame
                
    except StopIteration: pass
    return {"psnr_curve": psnr_curve, "avg_check_points_curve": avg_check_points_curve, "processing_time_curve": processing_time_curve, "total_pairs": processed_frame_pairs}

# ==========================================
# 系統預熱
# ==========================================
def warmup_system(algorithms: Dict[str, Callable], search_range: int, executor: concurrent.futures.ThreadPoolExecutor):
    print("--- 系統預熱中 (Warming up)... 請稍候 ---")
    dummy_h, dummy_w = 64, 64
    dummy_cur = np.zeros((dummy_h, dummy_w), dtype=np.uint8)
    dummy_ref = np.zeros((dummy_h, dummy_w), dtype=np.uint8)
    for algo_name, algo_func in algorithms.items():
        future = executor.submit(
            process_row_of_blocks, 
            0, dummy_w, dummy_cur, dummy_ref, algo_func, search_range
        )
        _ = future.result()
    print("--- 預熱完成！開始正式分析 ---")

# ==========================================
# 繪圖函式
# ==========================================
def plot_results(all_results: Dict[str, Any], output_dir: str):
    if not os.path.exists(output_dir): os.makedirs(output_dir)
    import matplotlib.pyplot as plt
    
    for video_name, algos_data in all_results.items():
        if not algos_data: continue
        
        fig, (ax1, ax2, ax3) = plt.subplots(3, 1, figsize=(10, 16))
        
        ax1.set_title(f'{video_name} - PSNR (Quality)')
        ax1.set_ylabel('PSNR (dB)')
        
        ax2.set_title(f'{video_name} - Search Points (Complexity)')
        ax2.set_ylabel('Avg. Checks per MB')
        
        ax3.set_title(f'{video_name} - Execution Time')
        ax3.set_ylabel('Seconds per Frame')
        
        fs_summary_text = "Full Search (Benchmark) Averages:\n"
        has_fs = False

        for algo_name, stats in algos_data.items():
            avg_checks = np.mean(stats["avg_check_points_curve"])
            avg_time = np.mean(stats["processing_time_curve"])
            
            if "Full" in algo_name:
                has_fs = True
                fs_summary_text += f"  - Checks: {avg_checks:.1f}\n"
                fs_summary_text += f"  - Time:   {avg_time:.4f}s\n"
                ax1.plot(stats["psnr_curve"], label=algo_name, marker='o', markersize=3, linestyle='--', color='black', alpha=0.5)
            else:
                ax1.plot(stats["psnr_curve"], label=algo_name, marker='o', markersize=3)
                ax2.plot(stats["avg_check_points_curve"], label=algo_name, marker='x', markersize=3)
                ax3.plot(stats["processing_time_curve"], label=algo_name, marker='s', markersize=3)
        
        if has_fs:
            fig.text(0.5, 0.02, fs_summary_text, ha='center', fontsize=10, 
                     bbox=dict(boxstyle="round,pad=0.5", fc="white", ec="black", alpha=0.8))
            plt.subplots_adjust(bottom=0.15)

        for ax in [ax1, ax2, ax3]:
            ax.legend()
            ax.grid(True)
            ax.set_xlabel('Frame Index')
            
        fig.tight_layout(rect=[0, 0.15, 1, 1])
        save_path = os.path.join(output_dir, f"{video_name}_analysis.png")
        fig.savefig(save_path)
        plt.close(fig)

# ==========================================
# [新功能] 進階 Excel 匯出
# ==========================================
def save_results_to_excel(all_results: Dict[str, Any], output_path: str):
    if not openpyxl:
        print("尚未安裝 openpyxl，無法輸出 Excel。")
        return

    wb = openpyxl.Workbook()
    
    # --- Sheet 1: Summary ---
    ws_summary = wb.active
    ws_summary.title = "Summary"
    # 新增 FPS 與 Total Time
    ws_summary.append(["Video", "Algorithm", "Avg PSNR (dB)", "Avg Checks", "Avg Time/Frame (s)", "Est. FPS", "Total Time (s)", "Frames Processed"])
    
    for v_name, algos in all_results.items():
        for a_name, s in algos.items():
            avg_psnr = np.mean(s.get('psnr_curve', [0]))
            avg_checks = np.mean(s.get('avg_check_points_curve', [0]))
            avg_time = np.mean(s.get('processing_time_curve', [0]))
            total_frames = s.get('total_pairs', 0)
            total_time = sum(s.get('processing_time_curve', [0]))
            fps = 1.0 / avg_time if avg_time > 0 else 0
            
            ws_summary.append([
                v_name, a_name,
                round(avg_psnr, 4),
                round(avg_checks, 2),
                round(avg_time, 5),
                round(fps, 2),
                round(total_time, 4),
                total_frames
            ])

    # --- Sheet 2: Frame Details (新增功能) ---
    ws_details = wb.create_sheet("Frame_Data")
    ws_details.append(["Video", "Algorithm", "Frame Index", "PSNR (dB)", "Checks", "Time (s)"])
    
    for v_name, algos in all_results.items():
        for a_name, s in algos.items():
            psnr_c = s.get('psnr_curve', [])
            checks_c = s.get('avg_check_points_curve', [])
            time_c = s.get('processing_time_curve', [])
            
            for i in range(len(psnr_c)):
                ws_details.append([
                    v_name, a_name, i+1,
                    psnr_c[i], checks_c[i], time_c[i]
                ])

    wb.save(output_path)
    print(f"\n✅ 詳細 Excel 報告已儲存: {output_path}")


if __name__ == "__main__":
    
    run_timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_root = f"ME_Run_{run_timestamp}"
    os.makedirs(output_root, exist_ok=True)
    plots_dir = os.path.join(output_root, "plots")
    os.makedirs(plots_dir, exist_ok=True)
    
    max_workers = os.cpu_count() or 4
    
    print(f"=== 批次處理開始 (Loop Testing Mode) ===")
    print(f"=== 循環測試次數: {TEST_LOOPS} 次 (將取平均時間) ===")
    print(f"=== 偵測到 CPU 核心數: {max_workers} ===")

    VIDEO_FILES = {
        "Garden": r"video/garden_sif.y4m", 
        "Tennis": r"video/tennis_sif.y4m", 
        "Football": r"video/football_cif.y4m", 
    }
    
    ALGORITHMS_TO_RUN = {
        "FullSearch": full_search,
        "TSS": three_step_search,
        "Diamond": diamond_search,
        "HexagonBS": hexagon_search, 
    }
    
    all_results = {}
    
    with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
        
        warmup_system(ALGORITHMS_TO_RUN, SEARCH_RANGE, executor)
        
        for v_name, v_path in VIDEO_FILES.items():
            if not os.path.exists(v_path):
                print(f"⚠️ 找不到影片: {v_path}，已跳過。")
                continue
                
            print(f"\n>>> 正在處理影片: {v_name}")
            all_results[v_name] = {}
            
            for algo_name, algo_func in ALGORITHMS_TO_RUN.items():
                print(f"   執行演算法: {algo_name} ...", end="", flush=True)
                
                # --- [新功能] 循環測試平均邏輯 ---
                accumulated_time_curve = None
                final_psnr_curve = []
                final_checks_curve = []
                final_total_pairs = 0
                
                for loop_idx in range(TEST_LOOPS):
                    # 每次都要重新開檔
                    reader = Y4MReader(v_path)
                    iter_frames = itertools.islice(reader, MAX_FRAMES + 1) if MAX_FRAMES else reader
                    
                    stats = analyze_sequence(iter_frames, algo_func, SEARCH_RANGE, executor)
                    reader.close()
                    
                    # 第一次迴圈保留 PSNR 和 Checks 數據 (因為這些是固定的)
                    if loop_idx == 0:
                        final_psnr_curve = stats['psnr_curve']
                        final_checks_curve = stats['avg_check_points_curve']
                        final_total_pairs = stats['total_pairs']
                        accumulated_time_curve = np.array(stats['processing_time_curve'])
                    else:
                        # 後續迴圈累加時間
                        current_time_curve = np.array(stats['processing_time_curve'])
                        # 確保長度一致 (理論上會一致)
                        min_len = min(len(accumulated_time_curve), len(current_time_curve))
                        accumulated_time_curve = accumulated_time_curve[:min_len] + current_time_curve[:min_len]

                    print(f" {loop_idx+1}", end="", flush=True)
                
                # 計算平均時間
                avg_time_curve = (accumulated_time_curve / TEST_LOOPS).tolist()
                
                # 組合最終結果
                final_stats = {
                    "psnr_curve": final_psnr_curve,
                    "avg_check_points_curve": final_checks_curve,
                    "processing_time_curve": avg_time_curve,
                    "total_pairs": final_total_pairs
                }
                
                all_results[v_name][algo_name] = final_stats
                
                avg_time = np.mean(avg_time_curve) if avg_time_curve else 0
                print(f" | Avg Time: {avg_time:.4f}s")

    # 輸出結果
    if all_results:
        excel_path = os.path.join(output_root, f"ME_Result_Detailed_{run_timestamp}.xlsx")
        save_results_to_excel(all_results, excel_path)
        plot_results(all_results, plots_dir)
        print(f"✅ 圖表已儲存至: {plots_dir}")

    print("\n=== 全部完成 ===")