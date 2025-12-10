import numpy as np
from typing import IO, Iterator, Tuple, List, Dict, Optional, Callable, Any
import itertools
import os
import sys
import math
# from dataclasses import dataclass # <-- 移除
from PIL import Image
import matplotlib.pyplot as plt
import time

# --- [優化] 匯入 Numba JIT ---
# 需要先安裝: pip install numba
try:
    # [FIX] 將所有 numba 相關的 import 移入 try 區塊
    from numba import jit
    from numba.experimental import jitclass
    from numba import int32, int64

except ImportError:
    print("警告: 'numba' 函式庫未找到。")
    print("程式將以純 Python 模式執行，速度會慢很多。")
    print("請執行 'pip install numba' 來大幅加速。")
    
    # 建立一個假的裝飾器，使程式碼在沒有 numba 時也能執行
    def jit(nopython=True):
        def decorator(func):
            return func
        return decorator

    # [FIX] 建立一個假的 jitclass 裝裝飾器 (回傳原始 class)
    def jitclass(spec):
        def decorator(cls):
            return cls
        return decorator

    # [FIX] 建立假的型別 (使用 Python int 作為備援)
    int32 = int
    int64 = int
# --- [優化] 結束 ---

# --- [新功能] 匯入 openpyxl ---
# 需要先安裝: pip install openpyxl
try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    print("警告: 'openpyxl' 函式庫未找到。")
    print("將無法儲存為 Excel (.xlsx) 檔案。")
    print("請執行 'pip install openpyxl' 來啟用此功能。")
    openpyxl = None
# --- [新功能] 結束 ---

# --- [GPU] 匯入 Numba CUDA ---
try:
    from numba import cuda
    print("Numba CUDA (GPU 支援) 匯入成功。")
except ImportError:
    print("警告: Numba CUDA 匯入失敗。無法使用 GPU 執行。")
    print("請確保已安裝 NVIDIA 驅動程式與 CUDA Toolkit。")
    cuda = None
# --- [GPU] 結束 ---

# --- 核心類別 ---
class Frame:
    """代表一幀視訊畫面，包含 Y, Cb, Cr 分量 (4:2:0)。"""
    def __init__(self, y: np.ndarray, cb: np.ndarray, cr: np.ndarray):
        self.y = y
        self.cb = cb
        self.cr = cr
        self.height, self.width = y.shape
        self.chroma_height, self.chroma_width = cb.shape
        
    def __repr__(self):
        """回傳幀的簡要資訊。"""
        return (f"<Frame Y:{self.y.shape} Cb:{self.cb.shape} Cr:{self.cr.shape}>")

    def iter_y_macroblocks(self) -> Iterator[Tuple[int, int, np.ndarray]]:
        """
        按 16x16 迭代宏區塊 (僅 Y (亮度) 分量)。
        
        Yields:
            (r, c, y_block_16x16)
            - r, c: 宏區塊左上角的 row, col 座標
            - y_block_16x16: 該位置的 16x16 亮度區塊
        """
        mb_size = 16
        for r in range(0, self.height, mb_size):
            for c in range(0, self.width, mb_size):
                y_block = self.y[r:r+mb_size, c:c+mb_size]
                yield r, c, y_block

    def get_y_block(self, r: int, c: int) -> np.ndarray:
        """取得單一 16x16 Y 區塊"""
        return self.y[r:r+16, c:c+16]

    @staticmethod
    def create_blank(width, height):
        """建立一個空白幀 (Y=128, Cb=128, Cr=128)"""
        y = np.full((height, width), 128, dtype=np.uint8)
        cb = np.full((height // 2, width // 2), 128, dtype=np.uint8)
        cr = np.full((height // 2, width // 2), 128, dtype=np.uint8)
        return Frame(y, cb, cr)

    def put_y_macroblock(self, r: int, c: int, y_block_16x16: np.ndarray):
        """將 16x16 Y 區塊放回指定位置 (用於重建幀)"""
        self.y[r:r+16, c:c+16] = y_block_16x16

class Y4MReader:
    """
    解析 YUV4MPEG2 (Y4M) 檔案，並作為 Frame 迭代器。
    """
    def __init__(self, filepath: str):
        self.file = open(filepath, 'rb')
        self.width = 0
        self.height = 0
        self.y_size = 0
        self.c_size = 0
        self._parse_header()

    def _parse_header(self):
        """解析 Y4M 檔案標頭，取得寬度與高度。"""
        header_line = self.file.readline().decode('ascii').strip()
        if not header_line.startswith('YUV4MPEG2'):
            raise ValueError("檔案不是有效的 YUV4MPEG2 格式")
        
        params = header_line.split(' ')
        for p in params:
            if p.startswith('W'):
                self.width = int(p[1:])
            elif p.startswith('H'):
                self.height = int(p[1:])
        
        if self.width == 0 or self.height == 0:
            raise ValueError("未能在 Y4M 標頭中找到寬度或高度")
            
        self.y_size = self.width * self.height
        self.c_size = (self.width // 2) * (self.height // 2)
        print(f"Y4M 讀取器：{self.width}x{self.height}, 4:2:0 (Y={self.y_size}, C={self.c_size})")

    def __iter__(self) -> Iterator[Frame]:
        return self

    def __next__(self) -> Frame:
        """讀取下一幀，並回傳 Frame 物件。"""
        frame_header = self.file.readline()
        if not frame_header:
            raise StopIteration
            
        if not frame_header.startswith(b'FRAME'):
            raise IOError(f"預期的 FRAME 標頭未找到，卻收到: {frame_header}")
        
        try:
            y_data = self.file.read(self.y_size)
            cb_data = self.file.read(self.c_size)
            cr_data = self.file.read(self.c_size)
            
            if len(y_data) != self.y_size:
                raise StopIteration
                
            y = np.frombuffer(y_data, dtype=np.uint8).reshape((self.height, self.width))
            cb = np.frombuffer(cb_data, dtype=np.uint8).reshape((self.height // 2, self.width // 2))
            cr = np.frombuffer(cr_data, dtype=np.uint8).reshape((self.height // 2, self.width // 2))
            
            return Frame(y, cb, cr)
        except Exception as e:
            print(f"讀取影格時發生錯誤: {e}")
            raise StopIteration

    def close(self):
        self.file.close()

def pad_frame_to_macroblock_boundary(frame: Frame) -> Frame:
    """確保幀尺寸是 16 的倍數"""
    h, w = frame.y.shape
    mh, mw = 16, 16
    pad_h = (mh - h % mh) % mh
    pad_w = (mw - w % mw) % mw
    if pad_h == 0 and pad_w == 0:
        return frame
    y_padded = np.pad(frame.y, ((0, pad_h), (0, pad_w)), mode='edge')
    pad_h_c = pad_h // 2
    pad_w_c = pad_w // 2
    cb_padded = np.pad(frame.cb, ((0, pad_h_c), (0, pad_w_c)), mode='edge')
    cr_padded = np.pad(frame.cr, ((0, pad_h_c), (0, pad_w_c)), mode='edge')
    print(f"Frame padding: Y ({h}x{w}) -> ({y_padded.shape})")
    return Frame(y_padded, cb_padded, cr_padded)

# --- 作業 1: 視覺化輔助工具 ---

def save_frames_as_images(frame_iterator: Iterator[Frame], output_dir: str, max_frames: int = 10):
    """
    將 Y4M 幀還原為一系列 PNG 圖片，以便視覺化。
    只儲存 Y (亮度) 分量。
    max_frames = 0 或小於 0 代表儲存所有影格。
    """
    if Image is None:
        print("錯誤: 'save_frames_as_images' 需要 Pillow 函式庫。")
        return

    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
        print(f"建立目錄: {output_dir}")

    saved = 0
    # max_frames <= 0 表示不限制（全部）
    if max_frames is None or max_frames <= 0:
        iterator = frame_iterator
    else:
        iterator = itertools.islice(frame_iterator, max_frames)

    for i, frame in enumerate(iterator):
        # 只關心 Y 分量
        y_channel = frame.y
        img = Image.fromarray(y_channel, mode='L')  # 'L' = 8-bit grayscale

        output_path = os.path.join(output_dir, f"frame_{i:04d}.png")
        img.save(output_path)
        saved += 1

    print(f"已儲存 {saved} 幀 (Y分量) 圖片到 {output_dir}")

# --- 作業 2: PSNR 與 ME 演算法 ---

# --- [優化] Numba jitclass ---
#
# 為了讓 JIT 函式 (nopython=True) 能夠建立並回傳一個 class 物件，
# 我們需要使用 @jitclass 來取代 @dataclass，並明確定義其型別。

spec = [
    ('mv_r', int32),
    ('mv_c', int32),
    ('sad', int64),      # SAD (Sum of Absolute Differences) 是整數加總
    ('check_points', int32),
]

@jitclass(spec)
class MotionVectorResult:
    """用於儲存 ME 的結果 (Numba JIT Class)"""
    def __init__(self, mv_r: int, mv_c: int, sad: int, check_points: int):
        self.mv_r = mv_r
        self.mv_c = mv_c
        self.sad = sad
        self.check_points = check_points

# --- [CPU] JIT 函式 (CPU 版) ---

# --- [優化] Level 1 (NumPy) + Level 2 (JIT) ---
@jit(nopython=True)
def calculate_sad(block1: np.ndarray, block2: np.ndarray) -> np.int64:
    """
    計算兩個區塊的 SAD (Sum of Absolute Differences) - 優化版
    使用 int16 避免浮點數轉換，並正確處理負數
    """
    # 將 block (uint8) 轉換為 int16 以避免溢位
    diff = block1.astype(np.int16) - block2.astype(np.int16)
    return np.sum(np.abs(diff))

# --- [優化] Level 2 (JIT) ---
@jit(nopython=True)
def get_reference_block(ref_y: np.ndarray, mb_r: int, mb_c: int, mv_r: int, mv_c: int) -> np.ndarray:
    """
    安全地從參考幀中提取一個 16x16 區塊。
    處理邊界情況 (運動向量指向畫面外)。
    args:
        ref_y (np.ndarray): 參考幀的 Y 分量(frame.y)
        mb_r (int): 當前宏塊的行索引
        mb_c (int): 當前宏塊的列索引
        mv_r (int): 垂直運動向量
        mv_c (int): 水平運動向量
    returns:
        np.ndarray: 提取的 16x16 參考區塊
    """
    h, w = ref_y.shape
    mb_size = 16
    
    # 計算參考區塊的左上角座標
    ref_r_start = mb_r + mv_r
    ref_c_start = mb_c + mv_c
    
    # 邊界箝位 (Clipping)
    # [FIX] Numba JIT (nopython=True) 對純量使用 np.clip 時會出錯
    # 改用手動箝位 (clamping)
    
    r_max = h - mb_size
    if ref_r_start < 0:
        ref_r_start_clipped = 0
    elif ref_r_start > r_max:
        ref_r_start_clipped = r_max
    else:
        ref_r_start_clipped = ref_r_start

    c_max = w - mb_size
    if ref_c_start < 0:
        ref_c_start_clipped = 0
    elif ref_c_start > c_max:
        ref_c_start_clipped = c_max
    else:
        ref_c_start_clipped = ref_c_start
    
    # 提取
    ref_block = ref_y[
        ref_r_start_clipped : ref_r_start_clipped + mb_size,
        ref_c_start_clipped : ref_c_start_clipped + mb_size
    ]
    return ref_block

def calculate_psnr(frame_orig: Frame, frame_recon: Frame) -> float:
    """
    計算兩幀 Y 分量之間的 PSNR。
    """
    # MSE (Mean Squared Error)
    mse = np.mean((frame_orig.y.astype(np.float32) - frame_recon.y.astype(np.float32)) ** 2)
    
    if mse == 0:
        return float('inf') # 兩張圖完全相同
        
    max_pixel = 255.0
    psnr = 20 * math.log10(max_pixel) - 10 * math.log10(mse)
    return psnr

# --- [CPU] Motion Estimation 演算法 (CPU 版) ---

# --- [優化] Level 2 (JIT) ---
@jit(nopython=True)
def full_search(
    current_mb_y: np.ndarray,  # 16x16
    ref_frame_y: np.ndarray,   # HxW
    mb_r: int, 
    mb_c: int,
    search_range: int = 7
):
    """
    實作 Full Search (FS) 演算法 (JIT 加速版)
    args:
        current_mb_y (np.ndarray): 當前宏塊的 Y 分量 (16x16)
        ref_frame_y (np.ndarray): 參考幀的 Y 分量 (HxW)
        mb_r (int): 當前宏塊的行索引
        mb_c (int): 當前宏塊的列索引
        search_range (int): 搜尋範圍 (預設為 7)
    returns:
        MotionVectorResult: 最佳運動向量
    """
    best_sad = np.int64(999999999999) # [FIX] Numba JIT: 初始值改為 large int
    best_mv_r, best_mv_c = 0, 0
    check_points = 0
    mb_size = 16

    # 搜尋範圍: -7 到 +7 (包含 0)
    # Numba 會將這層 Python 迴圈轉譯為高速機器碼
    for r_vec in range(-search_range, search_range + 1):
        for c_vec in range(-search_range, search_range + 1):
            
            # 1. 取得參考區塊 (呼叫 JIT 函式)
            ref_block = get_reference_block(ref_frame_y, mb_r, mb_c, r_vec, c_vec)
            
            # 2. 計算 SAD (呼叫 JIT 函式)
            sad = calculate_sad(current_mb_y, ref_block)
            check_points += 1
            
            # 3. 更新最佳 MV
            if sad < best_sad:
                best_sad = sad
                best_mv_r = r_vec
                best_mv_c = c_vec
                
    return MotionVectorResult(best_mv_r, best_mv_c, best_sad, check_points)

# --- [優化] Level 2 (JIT) ---
@jit(nopython=True)
def three_step_search(
    current_mb_y: np.ndarray,
    ref_frame_y: np.ndarray,
    mb_r: int, 
    mb_c: int,
    search_range: int = 7
):
    
    check_points = 0
    # Numba 不支援 List[int]，但支援 np.array
    steps = np.array([4, 2, 1]) 
    center_r, center_c = 0, 0
    best_mv_r, best_mv_c = 0, 0
    best_sad = np.int64(999999999999) # [FIX] Numba JIT: 初始值改為 large int

    # Numba 不支援 List[Tuple]，改用 np.array
    offsets_9_points = np.array([
        [0, 0], [0, 1], [0, -1], [1, 0], [-1, 0],
        [1, 1], [1, -1], [-1, 1], [-1, -1]
    ])
    
    for i in range(steps.shape[0]):
        step = steps[i]
        step_best_sad = np.int64(999999999999) # [FIX] Numba JIT: 初始值改為 large int
        step_best_r = center_r
        step_best_c = center_c
        
        # 決定要檢查哪些點
        if i == 0:
            # 第一步：檢查 9 個點（包含中心）
            points_to_check = offsets_9_points
        else:
            # 後續步驟：只檢查 8 個鄰居（不含中心）
            # Numba 的 array slicing
            points_to_check = offsets_9_points[1:]
        
        for j in range(points_to_check.shape[0]):
            dr_offset = points_to_check[j, 0]
            dc_offset = points_to_check[j, 1]
            
            mv_r = center_r + (dr_offset * step)
            mv_c = center_c + (dc_offset * step)

            # 邊界檢查
            if not (-search_range <= mv_r <= search_range and 
                    -search_range <= mv_c <= search_range):
                continue 

            ref_block = get_reference_block(ref_frame_y, mb_r, mb_c, mv_r, mv_c)
            sad = calculate_sad(current_mb_y, ref_block)
            check_points += 1
            
            # 更新本步驟最佳
            if sad < step_best_sad:
                step_best_sad = sad
                step_best_r = mv_r
                step_best_c = mv_c
        
        # 更新全局最佳
        if step_best_sad < best_sad:
            best_sad = step_best_sad
            best_mv_r = step_best_r
            best_mv_c = step_best_c
        
        # 移動中心到本步驟最佳位置
        center_r = step_best_r
        center_c = step_best_c

    return MotionVectorResult(best_mv_r, best_mv_c, best_sad, check_points)


# --- [注意] Diamond Search *不* 使用 JIT ---
# 因為 Numba (nopython=True) 不支援 Python 的 'set' 物件。
# 但它呼叫的 get_reference_block 和 calculate_sad 已經 JIT 編譯，
# 所以它依然會獲得 *顯著* 加速。
def diamond_search(
    current_mb_y: np.ndarray,
    ref_frame_y: np.ndarray,
    mb_r: int, 
    mb_c: int,
    search_range: int = 7
):
    """
    實作 Diamond Search (DS) 演算法
    
    使用 LDSP (大鑽石) 和 SDSP (小鑽石)
    檢查點數量是浮動的。
    """
    
    # LDSP (Large Diamond Search Pattern) - 9 點
    # (dx, dy) 相對於中心
    LDSP_OFFSETS = [
        (0, 0), (0, 2), (0, -2), (2, 0), (-2, 0),
        (1, 1), (1, -1), (-1, 1), (-1, -1)
    ]
    
    # SDSP (Small Diamond Search Pattern) - 5 點
    # (dx, dy) 相對於中心
    SDSP_OFFSETS = [
        (0, 0), (0, 1), (0, -1), (1, 0), (-1, 0)
    ]
    
    check_points = 0
    # 使用 set 來快取已檢查過的點 (MV)，避免重複計算
    checked_points_cache = set()

    # --- 步驟 1: 檢查 (0, 0) ---
    center_r, center_c = 0, 0
    
    # (呼叫 JIT 函式)
    ref_block = get_reference_block(ref_frame_y, mb_r, mb_c, center_r, center_c)
    # (呼叫 JIT 函式)
    best_sad = calculate_sad(current_mb_y, ref_block)
    best_mv_r, best_mv_c = center_r, center_c
    
    check_points = 1
    checked_points_cache.add((center_r, center_c))
    
    
    # --- 步驟 2: 迭代 LDSP 直到收斂 ---
    while True:
        current_center_r, current_center_c = best_mv_r, best_mv_c
        is_center_still_best = True
        
        # 檢查 LDSP (9 個點)
        for dr, dc in LDSP_OFFSETS:
            mv_r = current_center_r + dr
            mv_c = current_center_c + dc
            
            # --- 邊界檢查 ---
            if not (-search_range <= mv_r <= search_range and \
                    -search_range <= mv_c <= search_range):
                continue
            
            # --- 快取檢查 ---
            if (mv_r, mv_c) in checked_points_cache:
                continue
                
            # 檢查新點 (呼叫 JIT 函式)
            ref_block = get_reference_block(ref_frame_y, mb_r, mb_c, mv_r, mv_c)
            sad = calculate_sad(current_mb_y, ref_block)
            
            check_points += 1
            checked_points_cache.add((mv_r, mv_c))
            
            if sad < best_sad:
                best_sad = sad
                best_mv_r = mv_r
                best_mv_c = mv_c
                is_center_still_best = False # 找到了更好的點，中心移動
        
        # if 迴圈跑完，中心點 (current_center) 仍然是 LDSP 中的最佳點
        if is_center_still_best:
            break # LDSP 收斂，跳出 while 迴圈
        # else: best_mv 已更新，while 迴圈將以新中心繼續

    # --- 步驟 3: 最終 SDSP 檢查 ---
    # 在收斂的中心 (best_mv_r, best_mv_c) 周圍進行最後一次小範圍檢查
    
    # SDSP 的 5 個點
    for dr, dc in SDSP_OFFSETS:
        mv_r = best_mv_r + dr
        mv_c = best_mv_c + dc
        
        # --- 邊界檢查 ---
        if not (-search_range <= mv_r <= search_range and \
                -search_range <= mv_c <= search_range):
            continue
        
        # --- 快取檢查 ---
        if (mv_r, mv_c) in checked_points_cache:
            continue
            
        # 檢查新點 (呼叫 JIT 函式)
        ref_block = get_reference_block(ref_frame_y, mb_r, mb_c, mv_r, mv_c)
        sad = calculate_sad(current_mb_y, ref_block)
        
        check_points += 1
        checked_points_cache.add((mv_r, mv_c))
        
        # 這裡 *不* 更新 is_center_still_best，只更新最終的 best_mv
        if sad < best_sad:
            best_sad = sad
            best_mv_r = mv_r
            best_mv_c = mv_c

    return MotionVectorResult(best_mv_r, best_mv_c, best_sad, check_points)


# --- [CPU] 主分析邏輯 (CPU 版) ---

def reconstruct_frame_from_mvs(
    ref_frame: Frame, 
    mv_results: List[Tuple[int, int, MotionVectorResult]]
) -> Frame:
    """
    根據運動向量，從參考幀重建一個預測幀。
    """
    recon_frame = Frame.create_blank(ref_frame.width, ref_frame.height)
    
    for r, c, mv_res in mv_results:
        # 取得參考區塊 (呼叫 JIT 函式)
        ref_block_y = get_reference_block(ref_frame.y, r, c, mv_res.mv_r, mv_res.mv_c)
        
        # 將重建的區塊放回 P-Frame
        recon_frame.put_y_macroblock(r, c, ref_block_y)
        
    # 注意：為了 PSNR，我們只重建了 Y 分量。
    # 一個完整的 P-Frame 也會對 Cb, Cr 進行運動補償 (通常使用 Y 的 MV)
    # 但對於作業的 PSNR (Y) 測量，這樣足夠了。
    
    return recon_frame

def analyze_sequence(
    frame_iterator: Iterator[Frame], 
    algorithm_func: Callable, 
    search_range: int
) -> Dict[str, Any]:
    """
    執行 ME 分析的主迴圈。
    args:
        frame_iterator (Iterator[Frame]): Y4M 幀迭代器
        algorithm_func (Callable): ME 演算法函式 (full_search, three_step_search, diamond_search)
        search_range (int): 搜尋範圍
    returns:
        Dict[str, Any]: 包含 PSNR 曲線、平均檢查點曲線、處理時間曲線等結果的字典
    """
    
    # 儲存每幀結果的列表，用於繪製曲線
    psnr_curve = []
    avg_check_points_curve = []
    processing_time_curve = []
    
    processed_frame_pairs = 0
    
    # [JIT 預熱]
    # Numba JIT 在第一次呼叫函式時需要一點時間編譯。
    # 我們可以在計時迴圈開始前，手動 "預熱" 一次 JIT 函式。
    if 'numba' in sys.modules:
        print("  ... 正在預熱 (JIT 編譯) 演算法 ...")
        try:
            _dummy_y = np.zeros((32, 32), dtype=np.uint8)
            _dummy_mb = np.zeros((16, 16), dtype=np.uint8)
            # 傳入一個假的範圍以觸發編譯
            algorithm_func(_dummy_mb, _dummy_y, 0, 0, 1) 
            print("  ... 預熱完成 ...")
        except Exception as e:
            print(f"  ... JIT 預熱時發生錯誤 (可忽略): {e} ...")
    
    # 工作流程 1 & 2: 讀取並遍歷幀
    try:
        # 取得 frame_0 作為第一幀的 reference
        ref_frame = pad_frame_to_macroblock_boundary(next(frame_iterator))
        
        while True:
            # 取得 frame_i (current)
            current_frame = pad_frame_to_macroblock_boundary(next(frame_iterator))
            
            print(f"  ... 處理幀對 (Ref: {processed_frame_pairs}, Curr: {processed_frame_pairs + 1})")
            
            start_time = time.perf_counter()    # 計時開始
            mv_results_for_frame = []           # 儲存此幀的所有 MV 結果
            total_check_points_in_frame = 0     # 此幀的總檢查點數
            
            # 對每個 16x16 block 執行 ME
            for r, c, current_mb_y in current_frame.iter_y_macroblocks():
                
                # 執行指定的 ME 演算法 (已 JIT 加速)
                mv_result = algorithm_func(
                    current_mb_y,
                    ref_frame.y,
                    r, c,
                    search_range
                )
                
                mv_results_for_frame.append( (r, c, mv_result) )
                total_check_points_in_frame += mv_result.check_points

            if not mv_results_for_frame:
                continue # 避免除以零

            # 重建幀
            recon_frame = reconstruct_frame_from_mvs(ref_frame, mv_results_for_frame)
            
            # 計算重建影像品質 (PSNR)
            # (calculate_psnr 內部已包含 MSE 邏輯)
            psnr = calculate_psnr(current_frame, recon_frame)
            psnr_curve.append(psnr)
            
            # 評估運算量 (Avg check points per MV for this frame)
            total_mvs_in_frame = len(mv_results_for_frame)
            avg_check_points_for_frame = total_check_points_in_frame / total_mvs_in_frame
            avg_check_points_curve.append(avg_check_points_for_frame)
            
            # 記錄耗時
            end_time = time.perf_counter()
            duration = end_time - start_time
            processing_time_curve.append(duration)

            processed_frame_pairs += 1
            ref_frame = current_frame # 目前幀成為下一輪的參考幀
            
    except StopIteration:
        print(f"... 已到達序列結尾。共處理 {processed_frame_pairs} 個幀對。")
    
    return {
        "psnr_curve": psnr_curve,
        "avg_check_points_curve": avg_check_points_curve,
        "processing_time_curve": processing_time_curve,
        "total_pairs": processed_frame_pairs
    }

# --- [GPU] CUDA 核心與設備函式 ---

if cuda: # 只有在成功匯入 cuda 時才定義
    
    # --- [GPU] 設備函式 (在 GPU 上執行, 被 Kernel 呼叫) ---
    @cuda.jit(device=True)
    def calculate_sad_gpu(block1, block2):
        """
        GPU 版本的 SAD (Numba device function)
        """
        sad_sum = 0
        for i in range(16):
            for j in range(16):
                # Numba CUDA 不支援 .astype()，但支援直接的型別轉換
                diff = np.int16(block1[i, j]) - np.int16(block2[i, j])
                sad_sum += abs(diff)
        return np.int64(sad_sum)

    @cuda.jit(device=True)
    def get_reference_block_gpu(ref_y, mb_r, mb_c, mv_r, mv_c, h, w):
        """
        GPU 版本的 get_reference_block (Numba device function)
        注意: 它不能回傳一個新的 array，而是回傳一個 view (切片)
        Numba CUDA kernel 不支援 np.ndarray 回傳
        
        更新：更穩健的方法是傳入一個預先分配的 16x16 陣列 (shared memory 或 local array)
        並手動填充它。為求簡單，我們先嘗試直接切片。
        
        再次更新：直接在 full_search_kernel 內部計算 SAD，
        避免複雜的陣列傳遞。
        """
        pass # 此函式邏輯將被內聯 (inline) 到 kernel 中

    # --- [GPU] 核心函式 (在 GPU 上執行, 由 CPU 啟動) ---
    @cuda.jit
    def full_search_kernel(current_frame_y, ref_frame_y, search_range, results_array):
        """
        Full Search 的 GPU Kernel。
        每個執行緒 (Thread) 負責一個 16x16 的宏區塊 (Macroblock)。
        
        args:
            current_frame_y (DeviceArray): 當前幀 (在 GPU 上)
            ref_frame_y (DeviceArray): 參考幀 (在 GPU 上)
            search_range (int): 搜尋範圍 (例如 15)
            results_array (DeviceArray): 輸出的結果陣列 (M, N, 4)
                                         (mv_r, mv_c, sad, check_points)
        """
        
        # 1. 取得此執行緒負責的宏區塊 (MB) 座標 (r, c)
        # 我們使用 2D Grid，一個 Thread Block 處理一個 MB
        mb_c_idx = cuda.blockIdx.x
        mb_r_idx = cuda.blockIdx.y
        
        # 取得總網格大小 (即 MB 的總數)
        grid_w = cuda.gridDim.x
        grid_h = cuda.gridDim.y

        if mb_r_idx >= grid_h or mb_c_idx >= grid_w:
            return # 超出範圍的執行緒不執行

        # 宏區塊的左上角像素座標
        mb_r = mb_r_idx * 16
        mb_c = mb_c_idx * 16
        mb_size = 16
        
        # 取得幀的尺寸
        h, w = ref_frame_y.shape

        # 2. 執行 Full Search 邏輯 (與 CPU 版相同)
        best_sad = np.int64(999999999999)
        best_mv_r = 0
        best_mv_c = 0
        check_points = 0
        
        # 為了效能，將當前 MB 載入到一個小的 local array
        # (Numba CUDA 會試圖將其最佳化到暫存器)
        current_mb_y = cuda.local.array((16, 16), dtype=np.uint8)
        for i in range(mb_size):
            for j in range(mb_size):
                if (mb_r + i) < h and (mb_c + j) < w:
                    current_mb_y[i, j] = current_frame_y[mb_r + i, mb_c + j]

        for r_vec in range(-search_range, search_range + 1):
            for c_vec in range(-search_range, search_range + 1):
                
                # --- 內聯 (inline) get_reference_block_gpu 邏輯 ---
                ref_r_start = mb_r + r_vec
                ref_c_start = mb_c + c_vec
                
                # 手動箝位 (clamping)
                r_max = h - mb_size
                if ref_r_start < 0:
                    ref_r_start_clipped = 0
                elif ref_r_start > r_max:
                    ref_r_start_clipped = r_max
                else:
                    ref_r_start_clipped = ref_r_start

                c_max = w - mb_size
                if ref_c_start < 0:
                    ref_c_start_clipped = 0
                elif ref_c_start > c_max:
                    ref_c_start_clipped = c_max
                else:
                    ref_c_start_clipped = ref_c_start
                # --- 內聯結束 ---
                
                # --- 內聯 (inline) calculate_sad_gpu 邏輯 ---
                sad_sum = 0
                for i in range(mb_size):
                    for j in range(mb_size):
                        # 從 ref_frame 切片
                        ref_pixel = ref_frame_y[ref_r_start_clipped + i, ref_c_start_clipped + j]
                        cur_pixel = current_mb_y[i, j]
                        
                        diff = np.int16(cur_pixel) - np.int16(ref_pixel)
                        sad_sum += abs(diff)
                # --- 內聯結束 ---
                
                sad = np.int64(sad_sum)
                check_points += 1
                
                if sad < best_sad:
                    best_sad = sad
                    best_mv_r = r_vec
                    best_mv_c = c_vec
        
        # 3. 將結果寫回「結果陣列」
        results_array[mb_r_idx, mb_c_idx, 0] = best_mv_r
        results_array[mb_r_idx, mb_c_idx, 1] = best_mv_c
        results_array[mb_r_idx, mb_c_idx, 2] = best_sad
        results_array[mb_r_idx, mb_c_idx, 3] = check_points

# --- [GPU] 主分析邏輯 (GPU 版) ---

def reconstruct_frame_from_gpu_results(
    ref_frame: Frame, 
    gpu_results: np.ndarray # (M, N, 4) 的結果陣列
) -> Frame:
    """
    根據 GPU 傳回的結果陣列，重建一個預測幀。
    """
    recon_frame = Frame.create_blank(ref_frame.width, ref_frame.height)
    
    num_mbs_y, num_mbs_x, _ = gpu_results.shape
    
    for mb_r_idx in range(num_mbs_y):
        for mb_c_idx in range(num_mbs_x):
            
            # 宏區塊的像素座標
            r = mb_r_idx * 16
            c = mb_c_idx * 16
            
            # 從結果陣列讀取 MV
            mv_res = gpu_results[mb_r_idx, mb_c_idx]
            mv_r = int(mv_res[0])
            mv_c = int(mv_res[1])
            
            # 取得參考區塊 (使用 CPU JIT 函式)
            ref_block_y = get_reference_block(ref_frame.y, r, c, mv_r, mv_c)
            
            # 將重建的區塊放回 P-Frame
            recon_frame.put_y_macroblock(r, c, ref_block_y)
            
    return recon_frame

def analyze_sequence_gpu(
    frame_iterator: Iterator[Frame], 
    algorithm_kernel: Callable, # GPU Kernel (例如 full_search_kernel)
    search_range: int
) -> Dict[str, Any]:
    """
    執行 ME 分析的主迴圈 (GPU 版本)。
    """
    psnr_curve = []
    avg_check_points_curve = []
    processing_time_curve = []
    processed_frame_pairs = 0
    
    try:
        # 取得 frame_0 作為第一幀的 reference
        ref_frame = pad_frame_to_macroblock_boundary(next(frame_iterator))
        # [GPU] 將第一幀傳到 GPU
        d_ref_frame_y = cuda.to_device(ref_frame.y)
        
        while True:
            # 取得 frame_i (current)
            current_frame = pad_frame_to_macroblock_boundary(next(frame_iterator))
            
            print(f"  ... [GPU] 處理幀對 (Ref: {processed_frame_pairs}, Curr: {processed_frame_pairs + 1})")
            
            start_time = time.perf_counter() # 計時開始 (包含記憶體複製)
            
            # 1. [GPU] 將當前幀傳到 GPU
            d_current_frame_y = cuda.to_device(current_frame.y)
            
            # 2. [GPU] 計算 Grid/Block 維度
            h, w = current_frame.y.shape
            num_mbs_x = w // 16
            num_mbs_y = h // 16
            
            blocks_per_grid = (num_mbs_x, num_mbs_y)
            threads_per_block = (1, 1) # 1 個執行緒處理 1 個 MB
            
            # 3. [GPU] 分配 GPU 上的結果記憶體 (M, N, 4)
            d_results = cuda.device_array((num_mbs_y, num_mbs_x, 4), dtype=np.int64)
            
            # 4. [GPU] 啟動 Kernel
            algorithm_kernel[blocks_per_grid, threads_per_block](
                d_current_frame_y, 
                d_ref_frame_y, 
                search_range, 
                d_results
            )
            
            # 5. [GPU] 等待 GPU 完成
            cuda.synchronize()
            
            # 6. [GPU] 將結果複製回 CPU
            h_results = d_results.copy_to_host()
            
            # 7. [GPU] 重建幀 (在 CPU 上)
            recon_frame = reconstruct_frame_from_gpu_results(ref_frame, h_results)
            
            # 記錄耗時 (包含記憶體複製 + Kernel 執行 + 重建)
            end_time = time.perf_counter()
            duration = end_time - start_time
            processing_time_curve.append(duration)

            # 8. 計算 PSNR (在 CPU 上)
            psnr = calculate_psnr(current_frame, recon_frame)
            psnr_curve.append(psnr)
            
            # 9. 計算檢查點 (從 h_results)
            # h_results[..., 3] 包含了所有 check_points
            avg_check_points_for_frame = np.mean(h_results[:, :, 3])
            avg_check_points_curve.append(avg_check_points_for_frame)

            processed_frame_pairs += 1
            
            # 釋放上一幀的 GPU 記憶體
            del d_ref_frame_y 
            # 目前幀成為下一輪的參考幀 (d_current_frame_y 變為 d_ref_frame_y)
            d_ref_frame_y = d_current_frame_y 
            
            ref_frame = current_frame # CPU 端的 ref_frame 也要更新
            
    except StopIteration:
        print(f"... [GPU] 已到達序列結尾。共處理 {processed_frame_pairs} 個幀對。")
    
    return {
        "psnr_curve": psnr_curve,
        "avg_check_points_curve": avg_check_points_curve,
        "processing_time_curve": processing_time_curve,
        "total_pairs": processed_frame_pairs
    }

# --- 繪製結果曲線圖 ---

def plot_results(all_results: Dict[str, Any], output_dir: str):
    """
    [修正] 使用 Matplotlib 將分析結果繪製成曲線圖並儲存。
    """
    if plt is None:
        print("錯誤: 'plot_results' 需要 Matplotlib 函式庫。跳過繪圖。")
        return

    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
        print(f"建立繪圖目錄: {output_dir}")

    for video_name, algos_data in all_results.items():
        if not algos_data:
            continue
            
        print(f"  -> 正在繪製 '{video_name}' 的分析圖表...")
        
        # 建立一個 3x1 的圖表 (上下排列)，並拉高畫布高度
        fig, (ax1, ax2, ax3) = plt.subplots(3, 1, figsize=(12, 15))
        
        # --- 子圖 1: PSNR 曲線 ---
        ax1.set_title(f'{video_name} - PSNR per Frame')
        ax1.set_xlabel('Frame Number (i)')
        ax1.set_ylabel('PSNR (dB)')
                
        # --- 子圖 2: Check Points 曲線 ---
        ax2.set_title(f'{video_name} - Avg. Check Points per Frame')
        ax2.set_xlabel('Frame Number (i)')
        ax2.set_ylabel('Avg. Check Points per MV')
        
        # --- 子圖 3: Processing Time 曲線 ---
        ax3.set_title(f'{video_name} - Processing Time per Frame')
        ax3.set_xlabel('Frame Number (i)')
        ax3.set_ylabel('Time (seconds)')


        for algo_name, stats in algos_data.items():
            # 繪製 PSNR 曲線
            if stats.get("psnr_curve"):
                ax1.plot(stats["psnr_curve"], label=algo_name, marker='o', markersize=2, linestyle='--')

            # 繪製檢查點曲線
            if stats.get("avg_check_points_curve"):
                ax2.plot(stats["avg_check_points_curve"], label=algo_name, marker='x', markersize=2, linestyle=':')
            
            # 繪製時間曲線
            if stats.get("processing_time_curve"):
                ax3.plot(stats["processing_time_curve"], label=algo_name, marker='s', markersize=2, linestyle='-')

        # PSNR 曲線的圖例和格線
        ax1.legend()
        ax1.grid(True)

        # 檢查點曲線的圖例和格線
        ax2.legend()
        ax2.grid(True)
        
        # 時間曲線的圖例和格線
        ax3.legend()
        ax3.grid(True)

        # 調整佈局並儲存
        fig.tight_layout()
        output_path = os.path.join(output_dir, f"{video_name}_analysis_plot.png")
        fig.savefig(output_path)
        plt.close(fig)

    print(f"所有分析圖表已儲存至: {output_dir}")

# --- 程式進入點 ---

if __name__ == "__main__":
    try:
        if cuda.is_available():
            cuda.select_device(0)  # <--- 關鍵！強制選定第 0 號顯卡
            print(">>> 顯卡已強制喚醒！")
    except Exception as e:
        print(f"顯卡喚醒失敗: {e}")
    # ---------------------------

    # ... 下面接你原本的程式碼 ...
    # *** 確保你有這三個檔案的路徑 ***
    VIDEO_FILES = {
        "Garden": r"video/garden_sif.y4m", 
        "Tennis": r"video/tennis_sif.y4m", 
        "Football": r"video/football_cif.y4m", 
    }
    
    # --- [CPU] 演算法 ---
    ALGORITHMS_CPU = {
        "FullSearch": full_search,
        "3-Step-Search": three_step_search,
        "DiamondSearch": diamond_search,
    }
    
    # --- [GPU] 演算法 ---
    ALGORITHMS_GPU = {}
    if cuda:
        ALGORITHMS_GPU = {
            "FullSearch_GPU": full_search_kernel,
            # "3-Step-Search_GPU": ... (尚未實作)
            # "DiamondSearch_GPU": ... (移植困難)
            "3-Step-Search": three_step_search,
            "DiamondSearch": diamond_search,
        }

    # --- [修改] 執行模式開關 ---
    # 設為 True 會嘗試使用 GPU (如果可用)
    # 設為 False 會強制使用 CPU (Numba JIT)
    USE_GPU = True
    
    SEARCH_RANGE = 15
    # 限制分析的幀數以加速測試，設為 None 則跑完整個影片
    MAX_FRAMES_TO_ANALYZE = None # 設為 10 可以快速測試
    
    # -----------------------------------------------
    # 任務 1: 還原 frames (視覺化)
    # -----------------------------------------------
    run_visualization = False # 設為 False 可跳過此步驟
    if run_visualization and Image is not None:
        try:
            print("--- 任務 1: 還原 Frames (視覺化) ---")
            viz_reader = Y4MReader(VIDEO_FILES["Tennis"])
            save_frames_as_images(viz_reader, output_dir="tennis_frames", max_frames=5)
            viz_reader.close()
            print("-------------------------------------------\n")
        except Exception as e:
            print(f"視覺化時發生錯誤: {e}")

    # -----------------------------------------------
    # 任務 2, 3, 4: 測量與比較
    # -----------------------------------------------
    print("--- 任務 2, 3, 4: ME 分析 ---")
    
    # 決定要執行的演算法
    if USE_GPU and cuda and ALGORITHMS_GPU:
        print("\n*** 將使用 GPU 模式執行 ***")
        ALGORITHMS_TO_RUN = ALGORITHMS_GPU
        analyze_func = analyze_sequence_gpu
    else:
        if USE_GPU and not cuda:
            print("\n*** 警告: USE_GPU=True 但 Numba CUDA 不可用，將退回 CPU 模式 ***")
        print("\n*** 將使用 CPU (Numba JIT) 模式執行 ***")
        ALGORITHMS_TO_RUN = ALGORITHMS_CPU
        analyze_func = analyze_sequence
    
    
    all_results: Dict[str, Any] = {} # 儲存所有結果 (包含曲線)
    
    for video_name, video_path in VIDEO_FILES.items():
        if not os.path.exists(video_path):
            print(f"找不到檔案: {video_path}，跳過 '{video_name}' 分析。")
            continue
            
        print(f"\n===== 正在分析影片: {video_name} =====")
        all_results[video_name] = {}
        
        for algo_name, algo_func in ALGORITHMS_TO_RUN.items(): # [修改] 使用 ALGORITHMS_TO_RUN
            print(f"--- 使用演算法: {algo_name} ---")
            
            reader = Y4MReader(video_path)
            
            # 限制幀數 (包含第一幀 ref_frame)
            if MAX_FRAMES_TO_ANALYZE is None or MAX_FRAMES_TO_ANALYZE <= 0:
                # +1 因為第一幀是參考幀，不算在 "幀對" 內
                limited_frames = reader
            else:
                limited_frames = itertools.islice(reader, MAX_FRAMES_TO_ANALYZE + 1)
            
            # stats 包含 'psnr_curve' 和 'avg_check_points_curve'
            # [修改] 使用 analyze_func (可能是 CPU 或 GPU 版本)
            stats = analyze_func(limited_frames, algo_func, SEARCH_RANGE) 
            
            # --- 從曲線計算總平均值---
            avg_psnr = np.mean(stats['psnr_curve']) if stats['psnr_curve'] else 0
            avg_check_points = np.mean(stats['avg_check_points_curve']) if stats['avg_check_points_curve'] else 0
            avg_time = np.mean(stats['processing_time_curve']) if stats['processing_time_curve'] else 0
            
            print(f"  結果 (處理 {stats['total_pairs']} 幀對):")
            print(f"  -> [總平均] PSNR: {avg_psnr:.2f} dB")
            print(f"  -> [總平均] 檢查點/MV: {avg_check_points:.2f}")
            print(f"  -> [總平均] 處理時間: {avg_time:.4f} s/frame")
            
            all_results[video_name][algo_name] = stats
            reader.close()

    # -----------------------------------------------
    # 任務 5: 結果比較 (表格) 並儲存到 Excel
    # -----------------------------------------------
    print("\n\n===== 任務 5: 最終結果摘要 (總平均) =====")
    
    # --- [新功能] 建立 Excel 活頁簿 ---
    if openpyxl:
        print("\n--- 正在建立 Excel 報告: analysis_results.xlsx ---")
        wb = openpyxl.Workbook()
        
        # 建立「摘要」工作表
        summary_ws = wb.active
        summary_ws.title = "Summary"
        
        summary_header = ['Video', 'Algorithm', 'Avg PSNR (dB)', 'Avg Check Points', 'Avg Time (s)']
        summary_ws.append(summary_header)
        
        # 調整摘要工作表的欄寬
        for i, col_title in enumerate(summary_header, 1):
            summary_ws.column_dimensions[get_column_letter(i)].width = max(len(col_title) + 2, 20)
    else:
        wb = None
        if 'openpyxl' not in sys.modules: # 只有在 import 失敗時才顯示
            print("\n--- 'openpyxl' 未安裝，跳過儲存 Excel 檔案 ---")
    # --- [新功能] 結束 ---

    print(f"{'Video':<10} | {'Algorithm':<15} | {'Avg PSNR (dB)':<15} | {'Avg Check Points':<15} | {'Avg Time (s)':<15}")
    print("-" * (10 + 15 + 15 + 15 + 15 + 13))
    
    for video_name, algos in all_results.items():
        for algo_name, stats in algos.items():
            avg_psnr = np.mean(stats['psnr_curve']) if stats.get('psnr_curve') else 0
            avg_check_points = np.mean(stats['avg_check_points_curve']) if stats.get('avg_check_points_curve') else 0
            avg_time = np.mean(stats['processing_time_curve']) if stats.get('processing_time_curve') else 0
            
            # (A) 印出到 Console
            print(f"{video_name:<10} | {algo_name:<15} | {avg_psnr:<15.2f} | {avg_check_points:<15.2f} | {avg_time:<15.4f}")

            # --- [新功能] 寫入 Excel ---
            if wb:
                try:
                    # (B) 寫入 "摘要" 工作表
                    summary_row_data = [
                        video_name,
                        algo_name,
                        avg_psnr,
                        avg_check_points,
                        avg_time
                    ]
                    summary_ws.append(summary_row_data)
                    
                    # (C) 建立並寫入 "逐幀" 工作表
                    frame_ws_title = f"{video_name}_{algo_name}"
                    # Excel 工作表標題長度限制為 31 個字元
                    frame_ws = wb.create_sheet(title=frame_ws_title[:31])
                    
                    frame_header = ['Frame', 'PSNR (dB)', 'Check Points', 'Time (s)']
                    frame_ws.append(frame_header)
                    
                    # 調整逐幀工作表的欄寬
                    for i, col_title in enumerate(frame_header, 1):
                        frame_ws.column_dimensions[get_column_letter(i)].width = max(len(col_title) + 2, 18)

                    # 取得曲線資料
                    psnr_c = stats.get('psnr_curve', [])
                    checks_c = stats.get('avg_check_points_curve', [])
                    time_c = stats.get('processing_time_curve', [])
                    
                    num_frames = len(psnr_c)
                    
                    for i in range(num_frames):
                        frame_ws.append([
                            i + 1, # 幀編號 (從 1 開始)
                            psnr_c[i] if i < len(psnr_c) else None,
                            checks_c[i] if i < len(checks_c) else None,
                            time_c[i] if i < len(time_c) else None
                        ])
                except Exception as e:
                    print(f"  警告: 寫入工作表 {algo_name} 時發生錯誤: {e}")
            # --- [新功能] 結束 ---

    # --- [新功能] 儲存 Excel 檔案 ---
    if wb:
        try:
            output_excel_path = "analysis_results.xlsx"
            wb.save(output_excel_path)
            print(f"\n--- 成功儲存 Excel 報告到: {output_excel_path} ---")
        except Exception as e:
            print(f"\n--- 錯誤: 無法儲存 Excel 檔案: {e} ---")
    # --- [新功能] 結束 ---

    # -----------------------------------------------
    # 繪製曲線圖
    # -----------------------------------------------
    if plt:
        print("\n--- 繪製曲線圖 ---")
        plot_results(all_results, output_dir="analysis_plots")