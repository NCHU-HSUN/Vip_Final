import concurrent.futures
import json
import math
import os
import sys
import time
from dataclasses import dataclass, field
from datetime import datetime
from typing import List, Tuple

import numpy as np

# ==========================================
# 1. ç³»çµ±è¨­å®šèˆ‡ JIT åŒ¯å…¥
# ==========================================

DEFAULT_SEARCH_RANGE = 32  # å°æ‡‰è«–æ–‡ Stage-1 çš„ Â±32 æœå°‹ç¯„åœ
HW_TRACE_SAMPLE_MBS = [(0, 0), (0, 16), (16, 0)]  # è§€å¯Ÿç¡¬é«” FSM æ˜ å°„
HEXBS_LHP = ((2, 0), (1, 2), (-1, 2), (-2, 0), (-1, -2), (1, -2))
HEXBS_SHP = ((1, 0), (0, 1), (-1, 0), (0, -1))

# åŒ¯å…¥ Numba åŠ é€Ÿ
try:
    from numba import jit, int32, int64
    from numba.experimental import jitclass
except ImportError:
    print("è­¦å‘Š: 'numba' æœªå®‰è£ã€‚ç¨‹å¼å°‡ä»¥ç´” Python æ…¢é€Ÿæ¨¡å¼åŸ·è¡Œã€‚")
    def jit(nopython=True, nogil=True):
        def decorator(func): return func
        return decorator
    def jitclass(cls_or_spec=None, spec=None):
        def decorator(cls): return cls
        return decorator
    int32 = int
    int64 = int

# åŒ¯å…¥ Excel è™•ç†
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None
    print("è­¦å‘Š: 'openpyxl' æœªå®‰è£ï¼Œå°‡ç„¡æ³•è¼¸å‡º Excel å ±è¡¨ã€‚")

# åŒ¯å…¥åœ–ç‰‡è™•ç†
try:
    from PIL import Image
except ImportError:
    print("è­¦å‘Š: æœªå®‰è£ Pillowï¼Œç„¡æ³•è¼¸å‡ºé©—è­‰åœ–ç‰‡ã€‚")
    Image = None

# ==========================================
# 2. æ ¸å¿ƒé¡åˆ¥å®šç¾©
# ==========================================

class Frame:
    def __init__(self, y: np.ndarray, cb: np.ndarray, cr: np.ndarray):
        self.y, self.cb, self.cr = y, cb, cr
        self.height, self.width = y.shape
        
    @staticmethod
    def create_blank(width, height):
        return Frame(np.full((height, width), 128, dtype=np.uint8),
                     np.full((height//2, width//2), 128, dtype=np.uint8),
                     np.full((height//2, width//2), 128, dtype=np.uint8))
                     
    def put_y_macroblock(self, r, c, block):
        self.y[r:r+16, c:c+16] = block

class Y4MReader:
    def __init__(self, filepath: str):
        if not os.path.exists(filepath):
            raise FileNotFoundError(f"æ‰¾ä¸åˆ°å½±ç‰‡æª”æ¡ˆ: {filepath}")
        self.file = open(filepath, 'rb')
        self._parse_header()

    def _parse_header(self):
        line = self.file.readline().decode('ascii').strip()
        for p in line.split(' '):
            if p.startswith('W'): self.width = int(p[1:])
            elif p.startswith('H'): self.height = int(p[1:])
        self.y_size = self.width * self.height
        self.c_size = (self.width // 2) * (self.height // 2)

    def __iter__(self): return self

    def __next__(self) -> Frame:
        if not self.file.readline(): raise StopIteration
        y = np.frombuffer(self.file.read(self.y_size), dtype=np.uint8).reshape((self.height, self.width))
        c_sz = self.c_size
        cb = np.frombuffer(self.file.read(c_sz), dtype=np.uint8).reshape((self.height//2, self.width//2))
        cr = np.frombuffer(self.file.read(c_sz), dtype=np.uint8).reshape((self.height//2, self.width//2))
        return Frame(y, cb, cr)

    def close(self): self.file.close()

def pad_frame(frame: Frame) -> Frame:
    h, w = frame.y.shape
    ph, pw = (16 - h % 16) % 16, (16 - w % 16) % 16
    if ph == 0 and pw == 0: return frame
    return Frame(np.pad(frame.y, ((0, ph), (0, pw)), 'edge'),
                 np.pad(frame.cb, ((0, ph//2), (0, pw//2)), 'edge'),
                 np.pad(frame.cr, ((0, ph//2), (0, pw//2)), 'edge'))

# ==========================================
# 3. JIT æ¼”ç®—æ³•èˆ‡è¨ˆç®—å‡½å¼
# ==========================================
spec = [('mv_r', int32), ('mv_c', int32), ('sad', int64), ('check_points', int32)]
@jitclass(spec)
class MotionVectorResult:
    def __init__(self, mv_r, mv_c, sad, check_points):
        self.mv_r, self.mv_c, self.sad, self.check_points = mv_r, mv_c, sad, check_points


@dataclass
class HexbsCandidateLog:
    stage: str
    center: Tuple[int, int]
    offset: Tuple[int, int]
    candidate: Tuple[int, int]
    sad: int


@dataclass
class HexbsMacroblockTrace:
    video: str
    frame_idx: int
    mb_row: int
    mb_col: int
    search_range: int
    centers: List[Tuple[int, int]] = field(default_factory=list)
    candidates: List[HexbsCandidateLog] = field(default_factory=list)
    best_vector: Tuple[int, int] = (0, 0)
    min_sad: int = 0

    def to_dict(self) -> dict:
        return {
            "video": self.video,
            "frame_idx": self.frame_idx,
            "macroblock": {"row": self.mb_row, "col": self.mb_col},
            "search_range": self.search_range,
            "centers": self.centers,
            "candidates": [
                {
                    "stage": c.stage,
                    "center": c.center,
                    "offset": c.offset,
                    "candidate": c.candidate,
                    "sad": c.sad,
                }
                for c in self.candidates
            ],
            "best_vector": self.best_vector,
            "min_sad": self.min_sad,
        }

@jit(nopython=True, nogil=True) 
def calculate_sad(b1: np.ndarray, b2: np.ndarray) -> int:
    return np.sum(np.abs(b1.astype(np.int16) - b2.astype(np.int16)))

@jit(nopython=True, nogil=True) 
def get_ref_block(ref: np.ndarray, r: int, c: int, dr: int, dc: int) -> np.ndarray:
    h, w = ref.shape
    rr = min(max(r + dr, 0), h - 16)
    cc = min(max(c + dc, 0), w - 16)
    return ref[rr:rr+16, cc:cc+16]

def calc_psnr(orig: Frame, recon: Frame) -> float:
    mse = np.mean((orig.y.astype(np.float32) - recon.y.astype(np.float32)) ** 2)
    return 20 * math.log10(255.0) - 10 * math.log10(mse) if mse > 0 else float('inf')

# --- æ¼”ç®—æ³• ---

@jit(nopython=True, nogil=True)
def algo_full_search(cur, ref, r, c, rng):
    best_sad, best_dr, best_dc, checks = 999999, 0, 0, 0
    for dr in range(-rng, rng + 1):
        for dc in range(-rng, rng + 1):
            sad = calculate_sad(cur, get_ref_block(ref, r, c, dr, dc))
            checks += 1
            if sad < best_sad: best_sad, best_dr, best_dc = sad, dr, dc
    return MotionVectorResult(best_dr, best_dc, best_sad, checks)

@jit(nopython=True, nogil=True)
def algo_tss(cur, ref, r, c, rng):
    best_sad, best_dr, best_dc, checks, step = 999999, 0, 0, 0, 4
    neighbors = [(-1, -1), (-1, 0), (-1, 1), (0, -1), (0, 1), (1, -1), (1, 0), (1, 1)]
    while step >= 1:
        check_center = (step == 4) 
        ls, lr, lc = best_sad, best_dr, best_dc
        if check_center:
            dr, dc = best_dr, best_dc
            if -rng <= dr <= rng and -rng <= dc <= rng:
                sad = calculate_sad(cur, get_ref_block(ref, r, c, dr, dc))
                checks += 1
                if sad < ls: ls, lr, lc = sad, dr, dc
        for i, j in neighbors:
            dr, dc = best_dr + i * step, best_dc + j * step
            if not (-rng <= dr <= rng and -rng <= dc <= rng): continue
            sad = calculate_sad(cur, get_ref_block(ref, r, c, dr, dc))
            checks += 1
            if sad < ls: ls, lr, lc = sad, dr, dc
        best_sad, best_dr, best_dc, step = ls, lr, lc, step // 2
    return MotionVectorResult(best_dr, best_dc, best_sad, checks)

@jit(nopython=True, nogil=True)
def algo_diamond(cur, ref, r, c, rng):
    LDSP = [(0,0), (0,2), (0,-2), (2,0), (-2,0), (1,1), (1,-1), (-1,1), (-1,-1)]
    SDSP = [(0,0), (0,1), (0,-1), (1,0), (-1,0)]
    cache = np.zeros((2*rng+1, 2*rng+1), dtype=np.int8)
    checks, best_dr, best_dc = 0, 0, 0
    best_sad = calculate_sad(cur, get_ref_block(ref, r, c, 0, 0))
    cache[rng, rng] = 1; checks += 1
    
    while True:
        ctr_r, ctr_c, found = best_dr, best_dc, False
        for i, j in LDSP:
            dr, dc = ctr_r + i, ctr_c + j
            if not (-rng<=dr<=rng and -rng<=dc<=rng) or cache[dr+rng, dc+rng]: continue
            sad = calculate_sad(cur, get_ref_block(ref, r, c, dr, dc))
            checks += 1; cache[dr+rng, dc+rng] = 1
            if sad < best_sad: best_sad, best_dr, best_dc, found = sad, dr, dc, True
        if not found: break
    for i, j in SDSP:
        dr, dc = best_dr + i, best_dc + j
        if not (-rng<=dr<=rng and -rng<=dc<=rng) or cache[dr+rng, dc+rng]: continue
        sad = calculate_sad(cur, get_ref_block(ref, r, c, dr, dc))
        checks += 1; cache[dr+rng, dc+rng] = 1
        if sad < best_sad: best_sad, best_dr, best_dc = sad, dr, dc
    return MotionVectorResult(best_dr, best_dc, best_sad, checks)

@jit(nopython=True, nogil=True)
def algo_hexbs(cur, ref, r, c, rng):
    cache = np.zeros((2*rng+1, 2*rng+1), dtype=np.int8)
    checks, best_dr, best_dc = 0, 0, 0
    best_sad = calculate_sad(cur, get_ref_block(ref, r, c, 0, 0))
    cache[rng, rng] = 1; checks += 1
    
    while True:
        ctr_r, ctr_c, center_best = best_dr, best_dc, True
        for i, j in HEXBS_LHP:
            dr, dc = ctr_r + i, ctr_c + j
            if not (-rng<=dr<=rng and -rng<=dc<=rng) or cache[dr+rng, dc+rng]: continue
            sad = calculate_sad(cur, get_ref_block(ref, r, c, dr, dc))
            checks += 1; cache[dr+rng, dc+rng] = 1
            if sad < best_sad: best_sad, best_dr, best_dc, center_best = sad, dr, dc, False
        if center_best: break
    for i, j in HEXBS_SHP:
        dr, dc = best_dr + i, best_dc + j
        if not (-rng<=dr<=rng and -rng<=dc<=rng) or cache[dr+rng, dc+rng]: continue
        sad = calculate_sad(cur, get_ref_block(ref, r, c, dr, dc))
        checks += 1; cache[dr+rng, dc+rng] = 1
        if sad < best_sad: best_sad, best_dr, best_dc = sad, dr, dc
    return MotionVectorResult(best_dr, best_dc, best_sad, checks)


def hexbs_hw_trace(video_name: str, frame_idx: int, mb_row: int, mb_col: int,
                   cur_frame: Frame, ref_frame: Frame, rng: int) -> HexbsMacroblockTrace:
    trace = HexbsMacroblockTrace(video=video_name, frame_idx=frame_idx,
                                 mb_row=mb_row, mb_col=mb_col, search_range=rng)
    cur_block = cur_frame.y[mb_row:mb_row+16, mb_col:mb_col+16]
    best_sad = calculate_sad(cur_block, get_ref_block(ref_frame.y, mb_row, mb_col, 0, 0))
    best_dr = 0
    best_dc = 0
    cache = {(0, 0)}
    trace.centers.append((0, 0))

    while True:
        ctr_r, ctr_c = best_dr, best_dc
        center_best = True
        for off_r, off_c in HEXBS_LHP:
            dr, dc = ctr_r + off_r, ctr_c + off_c
            if not (-rng <= dr <= rng and -rng <= dc <= rng) or (dr, dc) in cache:
                continue
            sad = int(calculate_sad(cur_block, get_ref_block(ref_frame.y, mb_row, mb_col, dr, dc)))
            cache.add((dr, dc))
            trace.candidates.append(HexbsCandidateLog(stage="LHP",
                                                      center=(ctr_r, ctr_c),
                                                      offset=(off_r, off_c),
                                                      candidate=(dr, dc),
                                                      sad=sad))
            if sad < best_sad:
                best_sad, best_dr, best_dc, center_best = sad, dr, dc, False
                trace.centers.append((dr, dc))
        if center_best:
            break

    for off_r, off_c in HEXBS_SHP:
        dr, dc = best_dr + off_r, best_dc + off_c
        if not (-rng <= dr <= rng and -rng <= dc <= rng) or (dr, dc) in cache:
            continue
        sad = int(calculate_sad(cur_block, get_ref_block(ref_frame.y, mb_row, mb_col, dr, dc)))
        cache.add((dr, dc))
        trace.candidates.append(HexbsCandidateLog(stage="SHP",
                                                  center=(best_dr, best_dc),
                                                  offset=(off_r, off_c),
                                                  candidate=(dr, dc),
                                                  sad=sad))
        if sad < best_sad:
            best_sad, best_dr, best_dc = sad, dr, dc
            trace.centers.append((dr, dc))

    trace.best_vector = (best_dr, best_dc)
    trace.min_sad = int(best_sad)
    return trace


def capture_hw_trace_samples(video_name: str, frame_idx: int, ref_frame: Frame, cur_frame: Frame,
                             samples: List[Tuple[int, int]], rng: int) -> List[dict]:
    records = []
    for row, col in samples:
        if row + 16 > cur_frame.height or col + 16 > cur_frame.width:
            continue
        trace = hexbs_hw_trace(video_name, frame_idx, row, col, cur_frame, ref_frame, rng)
        records.append(trace.to_dict())
    return records

# ==========================================
# 4. åˆ†ææµç¨‹ (Analysis Pipeline)
# ==========================================

def process_row_task(row_idx, width, cur_y, ref_y, algo_func, rng):
    results = []
    for col_idx in range(0, width, 16):
        cur_mb = cur_y[row_idx : row_idx+16, col_idx : col_idx+16].copy()
        res = algo_func(cur_mb, ref_y, row_idx, col_idx, rng)
        results.append(res)
    return row_idx, results

def save_verification_images(vname, algo_name, cur_frame, recon_frame, out_dir):
    if Image is None: return

    # ç¢ºä¿è¼¸å‡ºç›®éŒ„å­˜åœ¨
    out_dir = os.path.join(out_dir, "verification_images")
    if not os.path.exists(out_dir):
        try:
            os.makedirs(out_dir, exist_ok=True)
        except OSError:
            pass # é˜²æ­¢å¤šåŸ·è¡Œç·’ç«¶çˆ­

    try:
        recon_img = Image.fromarray(recon_frame.y)
        recon_img.save(f"{out_dir}/{vname}_{algo_name}_Recon.png")
        
        diff = np.abs(cur_frame.y.astype(int) - recon_frame.y.astype(int)).astype(np.uint8)
        diff_enhanced = diff * 5 
        diff_img = Image.fromarray(diff_enhanced)
        diff_img.save(f"{out_dir}/{vname}_{algo_name}_Error.png")
    except Exception as e:
        print(f"  [Error] å„²å­˜åœ–ç‰‡å¤±æ•—: {e}")

def analyze_video(name, path, algos, rng, executor, save_images=False, out_dir="", hw_trace_cfg=None):
    # æœ¬æ¬¡åŸ·è¡Œçš„çµ±è¨ˆæ•¸æ“š
    stats = {k: {'psnr':[], 'time':[], 'checks':[]} for k in algos}
    
    # è©³ç´°æ•¸æ“šè¨˜éŒ„ (Frame-by-Frame)
    # æ ¼å¼: [Algo, Frame_ID, PSNR, Checks, Time]
    detailed_data = []
    hw_trace_records = []
    
    try:
        reader = Y4MReader(path)
    except Exception as e:
        print(f"  [Error] ç„¡æ³•è®€å–å½±ç‰‡ {path}: {e}")
        return {}, []

    try:
        ref = pad_frame(next(reader))
    except StopIteration: return {}, []
        
    try:
        frame_idx = 0
        while True:
            try: cur = pad_frame(next(reader))
            except StopIteration: break
            
            for algo_name, algo_func in algos.items():
                t_start = time.perf_counter()
                futures = []
                for r in range(0, cur.height, 16):
                    futures.append(executor.submit(process_row_task, r, cur.width, cur.y, ref.y, algo_func, rng))
                
                total_checks = 0
                recon = Frame.create_blank(cur.width, cur.height)
                for f in concurrent.futures.as_completed(futures):
                    r_idx, row_res = f.result()
                    for i, mb_res in enumerate(row_res):
                        c_idx = i * 16
                        total_checks += mb_res.check_points
                        ref_blk = get_ref_block(ref.y, r_idx, c_idx, mb_res.mv_r, mb_res.mv_c)
                        recon.put_y_macroblock(r_idx, c_idx, ref_blk)
                
                t_end = time.perf_counter()
                mb_count = (cur.height // 16) * (cur.width // 16)
                
                frame_psnr = calc_psnr(cur, recon)
                frame_checks = total_checks / mb_count
                frame_time = t_end - t_start
                
                stats[algo_name]['time'].append(frame_time)
                stats[algo_name]['checks'].append(frame_checks)
                stats[algo_name]['psnr'].append(frame_psnr)
                
                detailed_data.append([algo_name, frame_idx, frame_psnr, frame_checks, frame_time])
                
                if save_images and frame_idx == 0:
                    save_verification_images(name, algo_name, cur, recon, out_dir)

            if hw_trace_cfg and frame_idx < hw_trace_cfg.get("max_frames", 1):
                samples = hw_trace_cfg.get("samples", HW_TRACE_SAMPLE_MBS)
                traces = capture_hw_trace_samples(name, frame_idx, ref, cur, samples, rng)
                hw_trace_records.extend(traces)
            
            ref = cur
            frame_idx += 1
    finally:
        reader.close()
    if hw_trace_records and out_dir:
        trace_file = hw_trace_cfg.get("filename", f"{name}_hw_mapping_trace.json") if hw_trace_cfg else f"{name}_hw_mapping_trace.json"
        trace_path = os.path.join(out_dir, trace_file)
        with open(trace_path, "w", encoding="utf-8") as f:
            json.dump(hw_trace_records, f, indent=2, ensure_ascii=False)
        print(f"  â†³ å·²è¼¸å‡ºç¡¬é«”æ˜ å°„è¿½è¹¤: {trace_path}")
    return stats, detailed_data, hw_trace_records

# ==========================================
# 5. å ±å‘Šç”Ÿæˆ (Ultimate ç‰ˆï¼šçµåˆ Detailed èˆ‡ Per Loop)
# ==========================================

def save_excel_report_ultimate(per_loop_data, frame_stats_avg, output_path):
    if not openpyxl: 
        print("æœªå®‰è£ openpyxlï¼Œç„¡æ³•è¼¸å‡º Excelã€‚")
        return
        
    wb = openpyxl.Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", fill_type="solid")
    
    # -----------------------------------------------------
    # Sheet 1: Summary (å¹³å‡èˆ‡æ¯”è¼ƒ - ä¾†è‡ª F ç‰ˆçš„é‚è¼¯)
    # -----------------------------------------------------
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum.append(["Video", "Algorithm", "Avg PSNR (dB)", "Avg Checks", "Avg Time (s)", "PSNR Diff (vs FS)", "Checks Red.%", "Speedup (vs FS)", "Total Samples (Loops)"])
    for cell in ws_sum[1]: cell.font = header_font; cell.fill = header_fill

    # èšåˆæ‰€æœ‰ Loop çš„æ•¸æ“šä¾†ç®—ç¸½å¹³å‡
    agg_data = {} # (Video, Algo) -> {'psnr':[], 'checks':[], 'time':[]}
    for record in per_loop_data:
        key = (record['video'], record['algo'])
        if key not in agg_data: agg_data[key] = {'psnr':[], 'checks':[], 'time':[]}
        agg_data[key]['psnr'].append(record['psnr'])
        agg_data[key]['checks'].append(record['checks'])
        agg_data[key]['time'].append(record['time'])
    
    # å–å¾— Full Search çš„åŸºæº–å€¼ (ç‚ºäº†æ¯”è¼ƒ)
    fs_benchmarks = {} # Video -> (psnr, checks, time)
    for (vid, algo), vals in agg_data.items():
        if algo == "FullSearch":
            fs_benchmarks[vid] = (np.mean(vals['psnr']), np.mean(vals['checks']), np.mean(vals['time']))

    # å¯«å…¥ç¸½è¡¨
    sorted_keys = sorted(agg_data.keys())
    for (vid, algo) in sorted_keys:
        vals = agg_data[(vid, algo)]
        avg_p, avg_c, avg_t = np.mean(vals['psnr']), np.mean(vals['checks']), np.mean(vals['time'])
        samples = len(vals['psnr'])
        
        p_diff, c_red, spd = "-", "-", "-"
        if vid in fs_benchmarks:
            fs_p, fs_c, fs_t = fs_benchmarks[vid]
            p_diff = f"{avg_p - fs_p:+.3f}"
            if fs_c > 0: c_red = f"{(fs_c - avg_c)/fs_c*100:.1f}%"
            if avg_t > 0: spd = f"{fs_t/avg_t:.2f}x"
            
        ws_sum.append([vid, algo, round(avg_p, 2), round(avg_c, 1), round(avg_t, 5), p_diff, c_red, spd, samples])
    
    # -----------------------------------------------------
    # Sheet 2: Per_Loop_Summary (æ¯è¼ªç´°ç¯€ - ä¾†è‡ª M ç‰ˆçš„é‚è¼¯)
    # -----------------------------------------------------
    ws_loop = wb.create_sheet(title="Per_Loop_Summary")
    ws_loop.append(["Loop ID", "Video", "Algorithm", "Avg PSNR", "Avg Checks", "Avg Time (s)"])
    for cell in ws_loop[1]: cell.font = header_font; cell.fill = header_fill
    
    for r in per_loop_data:
        ws_loop.append([r['loop_id'], r['video'], r['algo'], round(r['psnr'], 4), round(r['checks'], 1), round(r['time'], 6)])

    # -----------------------------------------------------
    # Sheet 3: Detailed Sheets (æ©«å‘å±•é–‹ - ä¾†è‡ª F ç‰ˆçš„é‚è¼¯)
    # -----------------------------------------------------
    videos = sorted(list(set(k[0] for k in frame_stats_avg.keys())))
    algos_set = sorted(list(set(k[1] for k in frame_stats_avg.keys())))
    if "FullSearch" in algos_set:
        algos_set.remove("FullSearch")
        algos_set.insert(0, "FullSearch")

    for vname in videos:
        ws = wb.create_sheet(title=f"{vname}_Details")
        headers = ["Frame Index"]
        for algo in algos_set:
            headers.extend([f"{algo}_PSNR", f"{algo}_Checks", f"{algo}_Time"])
        ws.append(headers)
        for cell in ws[1]: cell.font = header_font; cell.fill = header_fill
        
        max_frame = max([k[2] for k in frame_stats_avg.keys() if k[0] == vname])
        for f_idx in range(max_frame + 1):
            row_data = [f_idx]
            for algo in algos_set:
                key = (vname, algo, f_idx)
                if key in frame_stats_avg:
                    p, c, t = frame_stats_avg[key]
                    row_data.extend([round(p, 4), round(c, 1), round(t, 6)])
                else:
                    row_data.extend(["-", "-", "-"])
            ws.append(row_data)

    # è‡ªå‹•èª¿æ•´æ¬„å¯¬
    for sheet in wb.worksheets:
        for column in sheet.columns:
            max_length = 0
            col_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
                except: pass
            sheet.column_dimensions[col_letter].width = max_length + 2

    wb.save(output_path)
    print(f"\nâœ… çµ‚æ¥µç‰ˆ Excel å ±å‘Šå·²å„²å­˜: {output_path}")

def plot_charts(video_name, stats, out_dir):
    import matplotlib.pyplot as plt
    fig, (ax1, ax2, ax3) = plt.subplots(3, 1, figsize=(10, 16))
    ax1.set_title(f'{video_name} - PSNR'); ax1.set_ylabel('dB')
    ax2.set_title(f'{video_name} - Checks'); ax2.set_ylabel('Count')
    ax3.set_title(f'{video_name} - Time'); ax3.set_ylabel('Sec')
    
    fs_txt = "Full Search (Benchmark):\n"
    has_fs = False
    for name, data in stats.items():
        x = range(len(data['time']))
        if 'Full' in name:
            has_fs = True
            fs_txt += f" Checks:{np.mean(data['checks']):.1f}, Time:{np.mean(data['time']):.4f}s"
            ax1.plot(x, data['psnr'], 'k--', alpha=0.5, label=name)
        else:
            ax1.plot(x, data['psnr'], label=name); ax2.plot(x, data['checks'], label=name); ax3.plot(x, data['time'], label=name)
            
    for ax in [ax1, ax2, ax3]: ax.legend(); ax.grid(True)
    if has_fs: fig.text(0.5, 0.02, fs_txt, ha='center', bbox=dict(fc='white', alpha=0.9))
    fig.savefig(f"{out_dir}/{video_name}_analysis.png")
    plt.close(fig)

# ==========================================
# 6. ä¸»ç¨‹å¼
# ==========================================
if __name__ == '__main__':
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_dir = f"ME_Ultimate_{ts}"
    os.makedirs(out_dir, exist_ok=True)
    os.makedirs(f"{out_dir}/plots", exist_ok=True)
    search_range = int(os.environ.get("HEXBS_SEARCH_RANGE", DEFAULT_SEARCH_RANGE))
    
    # 1. æ™ºæ…§è·¯å¾‘æœå°‹
    base_video_path = "video"
    if not os.path.exists(base_video_path):
        possible_paths = ["../video", "./video", "nchu-hsun/vip_final/NCHU-HSUN-Vip_Final-84430069098edb7f242cd73eaae9c798ab3ce302/video"]
        for p in possible_paths:
            if os.path.exists(p):
                base_video_path = p
                print(f"ğŸ“ åµæ¸¬åˆ°å½±ç‰‡è³‡æ–™å¤¾ä½ç½®: {base_video_path}")
                break
    
    VIDEOS = {
        "Garden": os.path.join(base_video_path, "garden_sif.y4m"), 
        "Football": os.path.join(base_video_path, "football_cif.y4m"), 
        "Tennis": os.path.join(base_video_path, "tennis_sif.y4m")
    }
    
    VALID_VIDEOS = {}
    for k, v in VIDEOS.items():
        if os.path.exists(v): VALID_VIDEOS[k] = v
        else: print(f"âš ï¸ è­¦å‘Š: æ‰¾ä¸åˆ°æª”æ¡ˆ {v}ï¼Œå°‡è·³éæ­¤æ¸¬è©¦ã€‚")
    
    ALGOS = {"FullSearch": algo_full_search, "TSS": algo_tss, "Diamond": algo_diamond, "HEXBS": algo_hexbs}
    
    workers = os.cpu_count() or 4
    NUM_RUNS = 3 # è¨­å®šåŸ·è¡Œæ¬¡æ•¸
    
    print(f"=== å•Ÿå‹•çµ‚æ¥µç‰ˆé©—è­‰èˆ‡æ•ˆèƒ½æ¸¬è©¦ (Workers: {workers}) ===")
    print(f"ğŸ“ çµæœå°‡å„²å­˜æ–¼: {out_dir}")
    print(f"ğŸ”§ æœå°‹ç¯„åœè¨­å®š: Â±{search_range} pixels")
    
    # è³‡æ–™å®¹å™¨
    per_loop_records = [] # å­˜æ¯ä¸€è¼ªçš„æ‘˜è¦ (List of Dict)
    frame_stats_accumulator = {} # å­˜æ¯ä¸€å¹€çš„è©³ç´°æ•¸æ“š (Key -> List of values)
    final_avg_for_plot = {} # å­˜æœ€å¾Œå¹³å‡çµ¦ç¹ªåœ–ç”¨
    hw_trace_paths = {}

    with concurrent.futures.ThreadPoolExecutor(max_workers=workers) as executor:
        for run_i in range(NUM_RUNS):
            loop_id = run_i + 1
            print(f"\nğŸ“¢ Run {loop_id}/{NUM_RUNS} ...")
            save_imgs = (run_i == 0) # åªåœ¨ç¬¬ä¸€è¼ªå­˜åœ–ç‰‡
            
            for vname, vpath in VALID_VIDEOS.items():
                print(f"   -> {vname}...", end='', flush=True)
                hw_cfg = None
                if run_i == 0:
                    hw_cfg = {
                        "samples": HW_TRACE_SAMPLE_MBS,
                        "max_frames": 1,
                        "filename": f"{vname}_hw_mapping_trace.json"
                    }
                # åŸ·è¡Œåˆ†æ
                cur_res, cur_details, hw_trace = analyze_video(
                    vname, vpath, ALGOS, search_range, executor,
                    save_images=save_imgs, out_dir=out_dir, hw_trace_cfg=hw_cfg)
                print(" Done")
                if hw_trace and hw_cfg:
                    hw_trace_paths[vname] = os.path.join(out_dir, hw_cfg["filename"])
                
                # A. è™•ç† Per Loop Data (ä¾†è‡ª M ç‰ˆçš„æ¦‚å¿µ)
                for algo, d in cur_res.items():
                    record = {
                        'loop_id': loop_id,
                        'video': vname,
                        'algo': algo,
                        'psnr': np.mean(d['psnr']),
                        'checks': np.mean(d['checks']),
                        'time': np.mean(d['time'])
                    }
                    per_loop_records.append(record)
                
                # B. ç´¯ç© Detailed Frame Data (ä¾†è‡ª F ç‰ˆçš„æ¦‚å¿µ)
                for row in cur_details:
                    # row = [Algo, Frame, PSNR, Checks, Time]
                    algo, fr, p, c, t = row
                    key = (vname, algo, fr)
                    if key not in frame_stats_accumulator:
                        frame_stats_accumulator[key] = {'psnr':[], 'checks':[], 'time':[]}
                    frame_stats_accumulator[key]['psnr'].append(p)
                    frame_stats_accumulator[key]['checks'].append(c)
                    frame_stats_accumulator[key]['time'].append(t)

                # C. ç´¯ç©çµ¦ç¹ªåœ–ç”¨çš„è³‡æ–™ (ç´¯åŠ æ™‚é–“)
                if vname not in final_avg_for_plot:
                    final_avg_for_plot[vname] = {algo: {'psnr': d['psnr'], 'checks': d['checks'], 'time': d['time']} for algo, d in cur_res.items()}
                else:
                    for algo in ALGOS:
                        final_avg_for_plot[vname][algo]['time'] = [sum(x) for x in zip(final_avg_for_plot[vname][algo]['time'], cur_res[algo]['time'])]

    if not per_loop_records:
        print("\nâŒ æ²’æœ‰ä»»ä½•å½±ç‰‡è¢«æˆåŠŸåˆ†æã€‚")
    else:
        print("\nğŸ“Š è¨ˆç®—å¹³å‡å€¼ä¸¦ç”¢å‡ºå ±å‘Š...")
        
        # 1. è™•ç†ç¹ªåœ–ç”¨çš„å¹³å‡æ•¸æ“š
        for vname in final_avg_for_plot:
            for algo in final_avg_for_plot[vname]:
                # æ™‚é–“é™¤ä»¥ RUNS æ¬¡æ•¸å–å¹³å‡
                final_avg_for_plot[vname][algo]['time'] = [t/NUM_RUNS for t in final_avg_for_plot[vname][algo]['time']]
            plot_charts(vname, final_avg_for_plot[vname], f"{out_dir}/plots")
            
        # 2. è™•ç† Frame Stats å¹³å‡ (çµ¦ Excel Detailed Sheet)
        frame_stats_avg = {}
        for key, vals in frame_stats_accumulator.items():
            avg_p = np.mean(vals['psnr'])
            avg_c = np.mean(vals['checks'])
            avg_t = np.mean(vals['time'])
            frame_stats_avg[key] = (avg_p, avg_c, avg_t)

        # 3. è¼¸å‡ºåŒ…å«æ‰€æœ‰ç²¾è¯çš„ Excel
        excel_name = f"{out_dir}/ME_Ultimate_Result_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        save_excel_report_ultimate(per_loop_records, frame_stats_avg, excel_name)
        
        print("\n=== çµ‚æ¥µç‰ˆæ¸¬è©¦å®Œæˆï¼ ===")
        if hw_trace_paths:
            print("ğŸ” å·²åŒæ­¥è¼¸å‡º HEXBS ç¡¬é«”æ˜ å°„ç´€éŒ„ï¼š")
            for vid, path in hw_trace_paths.items():
                print(f"   - {vid}: {path}")
